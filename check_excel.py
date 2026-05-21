
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx')
print("シート一覧:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    headers = [cell.value for cell in ws[1]]
    print("ヘッダー:", headers)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(v is not None for v in row):
            print(f"行{i}: {row}")
