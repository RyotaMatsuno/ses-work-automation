import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

print("=== フラップテック 全行 ===")
ws = wb["フラップテック"]
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if any(row):
        # 主要列だけ出力
        print(f"  行{i}: {row[:10]}")

print("\n=== TERRA 全行（4月入場付近）===")
ws = wb["TERRA"]
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if any(row):
        print(f"  行{i}: {row[:8]}")
