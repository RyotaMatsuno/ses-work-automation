# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

print("=== 全シート 稼働中データ確認 ===\n", flush=True)

# TERRA
ws = wb["TERRA"]
print("【TERRA】", flush=True)
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[1] and "稼働中" in str(row[2] or ""):
        # 担当,区分,ステータス,氏名,参画,期間,案件,単価,サイト,勤怠,更新,業務,仕入先,仕入単価,粗利,TERRA請求,岡本払出,実入り,備考
        name = row[3]
        tanto = row[0] or ""
        kubun = row[1] or ""
        terra_req = row[15]  # TERRA請求額
        okamoto_out = row[16]  # 岡本払出
        jissyu = row[17]  # 実入り
        site = row[8]  # 支払サイト
        print(
            f"  {name} / {tanto}{kubun} / TERRA請求:{terra_req} / 岡本払出:{okamoto_out} / 実入り:{jissyu} / サイト:{site}",
            flush=True,
        )

print("\n【グレイスライン】", flush=True)
ws = wb["グレイスライン"]
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] and "稼働中" in str(row[0]):
        # ステータス,氏名,参画,期間,案件/上位,案件単価,仕入単価,粗利,GL請求額,実入り,サイト,...,木原分
        name = row[1]
        gl_req = row[8]  # GL請求額
        jissyu = row[9]  # 実入り
        site = row[10]  # サイト
        kihara = row[14]  # 木原分
        print(f"  {name} / GL請求:{gl_req} / 実入り:{jissyu} / 木原:{kihara} / サイト:{site}", flush=True)

print("\n【フラップテック】", flush=True)
ws = wb["フラップテック"]
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[1] and "稼働中" in str(row[1]):
        # 担当,ステータス,氏名,参画,期間,案件/上位,案件単価,仕入単価,粗利,FT請求額,岡本払出,実入り,サイト,...,木原分
        name = row[2]
        tanto = row[0] or ""
        ft_req = row[9]
        okamoto_out = row[10]
        jissyu = row[11]
        site = row[12]
        kihara = row[15]
        print(
            f"  {name} / {tanto} / FT請求:{ft_req} / 岡本払出:{okamoto_out} / 実入り:{jissyu} / 木原:{kihara} / サイト:{site}",
            flush=True,
        )

print("\n【直契約】", flush=True)
ws = wb["直契約"]
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] and "稼働中" in str(row[0]):
        name = row[1]
        req = row[2]
        okamoto_out = row[4]
        jissyu = row[5]
        site = row[3]
        print(f"  {name} / 請求:{req} / 岡本払出:{okamoto_out} / 実入り:{jissyu} / サイト:{site}", flush=True)
