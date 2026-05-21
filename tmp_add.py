
import openpyxl, shutil
from datetime import datetime

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
BAK = EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
shutil.copy2(EXCEL_PATH, BAK)

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["TERRA"]

# 合計行（行37）の直前に挿入
# 現在の最終データ行=36（吉田祥平）、合計行=37
# 36の次、37の前に2行追加
ws.insert_rows(37, 2)

# 鶴川 行37
ws.cell(37, 1).value = None          # 担当
ws.cell(37, 2).value = "P"           # 区分
ws.cell(37, 3).value = "稼働中"       # ステータス
ws.cell(37, 4).value = "鶴川"        # 氏名
ws.cell(37, 5).value = "2026/5"      # 参画時期
ws.cell(37, 6).value = "長期"        # 期間
ws.cell(37, 7).value = "（確認中）"  # 案件

# 斎藤 行38
ws.cell(38, 1).value = None
ws.cell(38, 2).value = "P"
ws.cell(38, 3).value = "稼働中"
ws.cell(38, 4).value = "斎藤"
ws.cell(38, 5).value = "2026/5"
ws.cell(38, 6).value = "長期"
ws.cell(38, 7).value = "（確認中）"

wb.save(EXCEL_PATH)
print("追加完了: 鶴川(行37) 斎藤(行38) → TERRAシート 稼働中 2026/5")
print(f"バックアップ: {BAK}")
