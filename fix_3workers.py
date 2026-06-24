# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_t = wb["TERRA"]

# 行マップ
terra_row_map = {}
for i, row in enumerate(ws_t.iter_rows(min_row=4, values_only=True), start=4):
    if row[3]:
        terra_row_map[str(row[3])] = i

# 川崎(新): TERRA請求なし（FT経由）
r = terra_row_map.get("川崎(新)")
if r:
    ws_t.cell(row=r, column=16, value="請求なし")
    ws_t.cell(row=r, column=18, value="FT経由のためTERRA請求なし")
    print("[OK] 川崎(新): TERRA請求なし（FT経由）", flush=True)

# 齋藤よしまさ: 岡本担当（TERRA請求15,000・岡本払出9,000）
r = terra_row_map.get("齋藤よしまさ")
if r:
    ws_t.cell(row=r, column=1, value="岡本")  # 担当
    ws_t.cell(row=r, column=16, value=15000)  # TERRA請求
    ws_t.cell(row=r, column=17, value=9000)  # 岡本払出
    ws_t.cell(row=r, column=18, value=None)
    print("[OK] 齋藤よしまさ: 岡本担当 TERRA請求15,000 岡本払出9,000", flush=True)

# 橋詰(新): 現在の状態確認
r = terra_row_map.get("橋詰(新)")
if r:
    row_vals = [ws_t.cell(row=r, column=c).value for c in range(1, 19)]
    print(f"橋詰(新) 現状: {row_vals}", flush=True)

wb.save(XLSX_PATH)
print("\n[DONE] 保存完了", flush=True)
