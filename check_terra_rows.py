# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_t = wb["TERRA"]

# 3行目と4行目を確認
print("3行目:", [ws_t.cell(row=3, column=i).value for i in range(1, 22)], flush=True)
print("4行目:", [ws_t.cell(row=4, column=i).value for i in range(1, 22)], flush=True)
print("5行目:", [ws_t.cell(row=5, column=i).value for i in range(1, 22)], flush=True)
