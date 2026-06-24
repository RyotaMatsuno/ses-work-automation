# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)

# TERRAシートで吉田祥平を探す
ws_terra = wb["TERRA"]
print("=== TERRAシート 吉田祥平 ===", flush=True)
found = False
for row in ws_terra.iter_rows(min_row=4, values_only=True):
    name = str(row[3] or "")
    if "吉田" in name and "祥" in name:
        print(f"行データ: {row[:20]}", flush=True)
        found = True
if not found:
    print("TERRAシートに吉田祥平なし", flush=True)

# 橋本奈緒の岡本払出ロジック確認
ws_ft = wb["フラップテック"]
print("\n=== 橋本奈緒 行詳細（数式確認）===", flush=True)
for i, row in enumerate(ws_ft.iter_rows(min_row=4), start=4):
    cell_name = ws_ft.cell(row=i, column=3)
    if cell_name.value and "橋本" in str(cell_name.value):
        for j in range(1, 18):
            cell = ws_ft.cell(row=i, column=j)
            if cell.value is not None:
                print(f"  列{j}: {cell.value}", flush=True)
        break

# 吉田祥平のFT行も確認
print("\n=== 吉田祥平 FT行詳細 ===", flush=True)
for i, row in enumerate(ws_ft.iter_rows(min_row=4), start=4):
    cell_name = ws_ft.cell(row=i, column=3)
    if cell_name.value and "吉田" in str(cell_name.value):
        for j in range(1, 18):
            cell = ws_ft.cell(row=i, column=j)
            if cell.value is not None:
                print(f"  列{j}: {cell.value}", flush=True)
        break
