import json
from datetime import date

import openpyxl
import requests
from dateutil.relativedelta import relativedelta

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
TOKEN_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\freee_auth\freee_token.json"  # 新パス
FREEE_BASE = "https://api.freee.co.jp/api/1"
COMPANY_ID = 11712776


def load_token():
    with open(TOKEN_PATH, encoding="utf-8") as f:
        d = json.load(f)
    # リフレッシュチェック
    import time

    saved_at = d.get("saved_at", 0)
    expires_in = d.get("expires_in", 21600)
    if int(time.time()) - saved_at > expires_in - 300:
        res = requests.post(
            "https://accounts.secure.freee.co.jp/public_api/token",
            data={
                "grant_type": "refresh_token",
                "client_id": "731109064351970",
                "client_secret": "6rbUbEgQ1i58C7O6Ndg8TQDDQcoO6w9EGkCt_HkWADe9klxnGoN1iNd-vlF0vqkqdVOJYi8nfkYNY9M9evkBJQ",
                "refresh_token": d["refresh_token"],
            },
        )
        if res.status_code == 200:
            d = res.json()
            d["saved_at"] = int(time.time())
            with open(TOKEN_PATH, "w", encoding="utf-8") as f:
                json.dump(d, f, indent=2, ensure_ascii=False)
    return d["access_token"]


def freee_headers():
    return {"Authorization": f"Bearer {load_token()}", "Content-Type": "application/json"}


PARTNER_MAP = {
    "株式会社TERRA": 91256138,
    "株式会社フラップテック": 113795090,
    "株式会社グレイスライン": 93251323,
}


def load_active_entries():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    entries = []

    # --- TERRA ---
    ws = wb["TERRA"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "担当" in str(row[0]):
            header_row = i
            break
    if header_row is not None:
        for row in rows[header_row + 1 :]:
            if not any(row):
                continue
            tantou = str(row[0] or "").strip()
            kubun = str(row[1] or "").strip()
            status = str(row[2] or "").strip()
            name = str(row[3] or "").strip()
            case = str(row[6] or "").strip()
            tanka = row[7]
            shiire = row[13]  # 請求単価（仕入）

            if "稼働中" not in status:
                continue
            if not name or name in ("NaN", "", "稼働中合計"):
                continue
            if "稼働前" in status:
                continue

            is_gl_ft = any(k in case for k in ["グレイスライン", "フラップテック", "GL", "FT"])

            try:
                tanka_int = int(tanka) if tanka else 0
            except:
                tanka_int = 0
            try:
                shiire_int = int(shiire) if shiire else 0
            except:
                shiire_int = 0
            profit = tanka_int - shiire_int

            if kubun == "P":
                if is_gl_ft:
                    continue  # 請求なし
                seikyu = 15000
                rule = "プロパー→15,000円固定"
            elif kubun == "BP":
                if tantou in ("TERRA折半",):
                    seikyu = int(profit * 0.50)
                    rule = "TERRA折半→粗利×50%"
                else:
                    seikyu = int(profit * 0.80)
                    rule = f"BP({tantou})→粗利×80%"
            else:
                seikyu = 15000
                rule = "不明→15,000円固定"

            if seikyu <= 0:
                continue

            entries.append(
                {
                    "partner": "株式会社TERRA",
                    "partner_id": 91256138,
                    "name": name,
                    "profit": profit,
                    "seikyu": seikyu,
                    "rule": rule,
                    "source": "TERRA",
                }
            )

    # --- フラップテック ---
    ws = wb["フラップテック"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "担当" in str(row[0]):
            header_row = i
            break
    if header_row is not None:
        for row in rows[header_row + 1 :]:
            if not any(row):
                continue
            tantou = str(row[0] or "").strip()
            status = str(row[1] or "").strip()
            name = str(row[2] or "").strip()
            tanka = row[6]  # 請求単価(上位から)
            shiire = row[7]  # 請求単価(下位へ)

            if "稼働中" not in status:
                continue
            if not name or name in ("NaN", "", "稼働中合計"):
                continue
            if "稼働前" in status:
                continue

            try:
                tanka_int = int(tanka) if tanka else 0
            except:
                tanka_int = 0
            try:
                shiire_int = int(shiire) if shiire else 0
            except:
                shiire_int = 0
            profit = tanka_int - shiire_int

            if tantou == "小坂折半":
                seikyu = int(profit * 0.48)
                rule = "小坂折半→粗利×48%"
            elif tantou in ("岡本折半", "岡本"):
                seikyu = int(profit * 0.68)
                rule = f"{tantou}→粗利×68%"
            else:
                seikyu = int(profit * 0.68)
                rule = "通常→粗利×68%"

            if seikyu <= 0:
                continue

            entries.append(
                {
                    "partner": "株式会社フラップテック",
                    "partner_id": 113795090,
                    "name": name,
                    "profit": profit,
                    "seikyu": seikyu,
                    "rule": rule,
                    "source": "FT",
                }
            )

    # --- グレイスライン ---
    ws = wb["グレイスライン"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "ステータス" in str(row[0]):
            header_row = i
            break
    if header_row is not None:
        for row in rows[header_row + 1 :]:
            if not any(row):
                continue
            status = str(row[0] or "").strip()
            name = str(row[1] or "").strip()
            tanka = row[5]  # 請求単価(上位から)
            shiire = row[6]  # 請求単価(下位へ)

            if "稼働中" not in status:
                continue
            if not name or name in ("NaN", "", "稼働中合計"):
                continue
            if "稼働前" in status:
                continue

            try:
                tanka_int = int(tanka) if tanka else 0
            except:
                tanka_int = 0
            try:
                shiire_int = int(shiire) if shiire else 0
            except:
                shiire_int = 0
            profit = tanka_int - shiire_int
            seikyu = int(profit * 0.60)

            if seikyu <= 0:
                continue

            entries.append(
                {
                    "partner": "株式会社グレイスライン",
                    "partner_id": 93251323,
                    "name": name,
                    "profit": profit,
                    "seikyu": seikyu,
                    "rule": "GL→粗利×60%",
                    "source": "GL",
                }
            )

    return entries


def create_invoice_draft(entry, issue_date, due_date, target_mon_str):
    payload = {
        "company_id": COMPANY_ID,
        "issue_date": issue_date.strftime("%Y-%m-%d"),
        "due_date": due_date.strftime("%Y-%m-%d"),
        "partner_id": entry["partner_id"],
        "invoice_status": "draft",
        "title": f"{target_mon_str}分 業務委託料（{entry['name']}様）",
        "description": f"[{entry['rule']}] 粗利: {entry['profit']:,}円",
        "invoice_lines": [
            {
                "name": f"業務委託料（{entry['name']}様）{target_mon_str}分",
                "quantity": 1,
                "unit_price": entry["seikyu"],
                "tax_code": 1,
                "type": "normal",
            }
        ],
    }
    res = requests.post(f"{FREEE_BASE}/invoices", headers=freee_headers(), json=payload)
    if res.status_code in (200, 201):
        inv = res.json()["invoice"]
        print(
            f"  OK | {entry['name']} / {entry['partner']} / {entry['seikyu']:,}円 | {entry['rule']} → #{inv['invoice_number']}"
        )
        return True
    else:
        print(f"  NG | {entry['name']} | {res.status_code}: {res.text[:300]}")
        return False


def run(target_month):
    issue_date = target_month.replace(day=1)
    # 支払いサイト45日 → 翌月末
    due_date = (issue_date + relativedelta(months=2)).replace(day=1) - relativedelta(days=1)
    target_mon_str = f"{target_month.year}年{target_month.month}月"

    print("=== freee請求書自動生成 ===")
    print(f"対象: {target_mon_str}分")
    print(f"請求日: {issue_date}  支払期限: {due_date}")
    print()

    entries = load_active_entries()

    # 会社別にまとめて表示
    from collections import defaultdict

    by_partner = defaultdict(list)
    for e in entries:
        by_partner[e["partner"]].append(e)

    total_seikyu = {}
    for partner, elist in by_partner.items():
        subtotal = sum(e["seikyu"] for e in elist)
        total_seikyu[partner] = subtotal
        print(f"【{partner}】 {len(elist)}件 合計{subtotal:,}円（税別）")
        for e in elist:
            print(f"    {e['name']} : {e['seikyu']:,}円 [{e['rule']}]")

    print()
    print(f"全体: {len(entries)}件 / {sum(total_seikyu.values()):,}円（税別）")
    print()

    ok = ng = 0
    for e in entries:
        if create_invoice_draft(e, issue_date, due_date, target_mon_str):
            ok += 1
        else:
            ng += 1

    print()
    print(f"=== 完了: 作成{ok}件 / エラー{ng}件 ===")
    print("→ https://secure.freee.co.jp/invoices")


# 4月分で実行
run(date(2026, 4, 1))
