import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# 5月分の稼働状況を確認 (freee_invoice_apr.pyと同じロジックで稼働中を抽出)
from collections import defaultdict


def extract_entries():
    entries = []

    # --- TERRA ---
    ws = wb["TERRA"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "担当" in str(row[0]):
            header_row = i
            break
    print(f"TERRA ヘッダー行: {header_row}")
    if header_row is not None:
        for row in rows[header_row + 1 :]:
            if not any(row):
                continue
            tantou = str(row[0] or "").strip()
            kubun = str(row[1] or "").strip()
            status = str(row[2] or "").strip()
            name = str(row[3] or "").strip()
            case = str(row[6] or "").strip()
            tanka = row[7]
            shiire = row[13]
            if "稼働中" not in status:
                continue
            if not name or name in ("NaN", "", "稼働中合計"):
                continue
            if "稼働前" in status:
                continue
            is_gl_ft = any(k in case for k in ["グレイスライン", "フラップテック", "GL", "FT"])
            try:
                tanka_int = int(tanka) if tanka else 0
            except:
                tanka_int = 0
            try:
                shiire_int = int(shiire) if shiire else 0
            except:
                shiire_int = 0
            profit = tanka_int - shiire_int
            if kubun == "P":
                if is_gl_ft:
                    continue
                seikyu = 15000
                rule = "プロパー→15,000円固定"
            elif kubun == "BP":
                if tantou in ("TERRA折半",):
                    seikyu = int(profit * 0.50)
                    rule = "TERRA折半→粗利×50%"
                else:
                    seikyu = int(profit * 0.80)
                    rule = f"BP({tantou})→粗利×80%"
            else:
                seikyu = 15000
                rule = "不明→15,000円固定"
            if seikyu <= 0:
                continue
            entries.append(
                {
                    "partner": "株式会社TERRA",
                    "name": name,
                    "profit": profit,
                    "seikyu": seikyu,
                    "rule": rule,
                    "source": "TERRA",
                }
            )

    # --- フラップテック ---
    ws = wb["フラップテック"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "担当" in str(row[0]):
            header_row = i
            break
    print(f"FT ヘッダー行: {header_row}")
    if header_row is not None:
        for row in rows[header_row + 1 :]:
            if not any(row):
                continue
            tantou = str(row[0] or "").strip()
            status = str(row[1] or "").strip()
            name = str(row[2] or "").strip()
            tanka = row[6]
            shiire = row[7]
            if "稼働中" not in status:
                continue
            if not name or name in ("NaN", "", "稼働中合計"):
                continue
            if "稼働前" in status:
                continue
            try:
                tanka_int = int(tanka) if tanka else 0
            except:
                tanka_int = 0
            try:
                shiire_int = int(shiire) if shiire else 0
            except:
                shiire_int = 0
            profit = tanka_int - shiire_int
            if tantou == "小坂折半":
                seikyu = int(profit * 0.48)
                rule = "小坂折半→粗利×48%"
            elif tantou in ("岡本折半", "岡本"):
                seikyu = int(profit * 0.68)
                rule = f"{tantou}→粗利×68%"
            else:
                seikyu = int(profit * 0.68)
                rule = "通常→粗利×68%"
            if seikyu <= 0:
                continue
            entries.append(
                {
                    "partner": "株式会社フラップテック",
                    "name": name,
                    "profit": profit,
                    "seikyu": seikyu,
                    "rule": rule,
                    "source": "FT",
                }
            )

    # --- グレイスライン ---
    ws = wb["グレイスライン"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "ステータス" in str(row[0]):
            header_row = i
            break
    print(f"GL ヘッダー行: {header_row}")
    if header_row is not None:
        for row in rows[header_row + 1 :]:
            if not any(row):
                continue
            status = str(row[0] or "").strip()
            name = str(row[1] or "").strip()
            tanka = row[5]
            shiire = row[6]
            if "稼働中" not in status:
                continue
            if not name or name in ("NaN", "", "稼働中合計"):
                continue
            if "稼働前" in status:
                continue
            try:
                tanka_int = int(tanka) if tanka else 0
            except:
                tanka_int = 0
            try:
                shiire_int = int(shiire) if shiire else 0
            except:
                shiire_int = 0
            profit = tanka_int - shiire_int
            seikyu = int(profit * 0.60)
            if seikyu <= 0:
                continue
            entries.append(
                {
                    "partner": "株式会社グレイスライン",
                    "name": name,
                    "profit": profit,
                    "seikyu": seikyu,
                    "rule": "GL→粗利×60%",
                    "source": "GL",
                }
            )

    return entries


entries = extract_entries()
by_partner = defaultdict(list)
for e in entries:
    by_partner[e["partner"]].append(e)

print()
print("=" * 60)
print("5月分 請求予定 一覧")
print("=" * 60)
total = 0
for partner, elist in by_partner.items():
    subtotal = sum(e["seikyu"] for e in elist)
    total += subtotal
    print(f"\n【{partner}】 {len(elist)}件  小計 {subtotal:,}円")
    for e in elist:
        print(f"  {e['name']:15s} {e['seikyu']:>8,}円  [{e['rule']}]  粗利:{e['profit']:>6,}円")
print()
print(f"合計: {len(entries)}件 / {total:,}円（税別）")
