import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# GL全列確認
ws = wb["グレイスライン"]
print("=== GL 全列ヘッダー ===")
for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
    print(row)

print("\n=== GL 稼働中データ ===")
for row in ws.iter_rows(min_row=1, values_only=True):
    if row[0] and "稼働中" in str(row[0]):
        print(row)
