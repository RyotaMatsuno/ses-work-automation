# -*- coding: utf-8 -*-
"""
auto_status_update.py
freee_invoice_v2.pyが請求書作成後に呼び出す
→ 請求書に含まれた人員を契約マスターExcelで「稼働中」に自動更新

使い方:
  from auto_status_update import update_status_after_invoice
  update_status_after_invoice(names=["山田太郎", "鈴木花子"])

または単体テスト:
  python freee/auto_status_update.py --test 山田太郎
"""

import sys

import openpyxl

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
TARGET_STATUS = "稼働中"

# シート別のヘッダー行インデックスと列マッピング
SHEET_CONFIG = {
    "TERRA": {"header_row": 3, "name_col": 3, "status_col": 2},
    "グレイスライン": {"header_row": 2, "name_col": 1, "status_col": 0},
    "フラップテック": {"header_row": 2, "name_col": 2, "status_col": 1},
}


def update_status_after_invoice(names: list, sheet_names: list = None, dry_run: bool = False):
    """
    指定した人名リストを契約マスターで「稼働中」に更新する。
    names: 更新対象の人名リスト
    sheet_names: 対象シート（省略時はSHEET_CONFIG全シート）
    dry_run: Trueなら書き込みせず確認だけ
    """
    if not names:
        print("[auto_status] 更新対象なし")
        return []

    wb = openpyxl.load_workbook(EXCEL_PATH)
    target_sheets = sheet_names if sheet_names else list(SHEET_CONFIG.keys())
    updated = []

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"[auto_status] {sheet_name}: シートなしスキップ")
            continue
        cfg = SHEET_CONFIG.get(sheet_name)
        if not cfg:
            print(f"[auto_status] {sheet_name}: 設定なしスキップ")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows())

        for row in rows[cfg["header_row"] + 1 :]:
            name_cell = row[cfg["name_col"]]
            status_cell = row[cfg["status_col"]]

            if not name_cell.value:
                continue
            name_str = str(name_cell.value).strip()

            if name_str in names:
                current = str(status_cell.value or "").strip()
                if current != TARGET_STATUS:
                    if not dry_run:
                        status_cell.value = TARGET_STATUS
                    updated.append(
                        {
                            "sheet": sheet_name,
                            "name": name_str,
                            "before": current,
                            "after": TARGET_STATUS,
                        }
                    )
                    print(
                        f"[auto_status] {'[DRY] ' if dry_run else ''}{sheet_name}/{name_str}: 「{current}」→「{TARGET_STATUS}」"
                    )
                else:
                    print(f"[auto_status] {sheet_name}/{name_str}: 既に{TARGET_STATUS}")

    if not dry_run and updated:
        wb.save(EXCEL_PATH)
        print(f"[auto_status] 保存完了: {len(updated)}件更新")
    elif dry_run:
        print(f"[auto_status] DRY RUN完了: {len(updated)}件が更新対象")
    else:
        print("[auto_status] 更新なし")

    return updated


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    is_test = "--test" in sys.argv
    dry = "--dry" in sys.argv or is_test

    if args:
        print(f"=== {'DRY RUN ' if dry else ''}更新対象: {args} ===")
        result = update_status_after_invoice(names=args, dry_run=dry)
        print(f"結果: {result}")
    else:
        print("使い方:")
        print("  python freee/auto_status_update.py 山田太郎 鈴木花子       # 実際に更新")
        print("  python freee/auto_status_update.py --dry 山田太郎          # ドライラン")
        print("  python freee/auto_status_update.py --test 山田太郎         # ドライラン(同上)")
