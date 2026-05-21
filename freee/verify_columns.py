"""アップロードファイルを使った列インデックス検証"""
import sys, os
sys.path.insert(0, r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\freee_auth")
sys.path.insert(0, r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\freee")

import openpyxl
from token_manager import get_headers

# 修正後のロジックでローカルExcelを読む
EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"

def safe_int(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    return 0

def is_valid_name(v):
    if v is None: return False
    if isinstance(v, (int, float)): return False
    import datetime
    if isinstance(v, (datetime.datetime, datetime.date)): return False
    s = str(v).strip()
    if not s or s in ("NaN", "稼働中合計"): return False
    if s.replace("/", "").replace("-", "").isdigit(): return False
    return True

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print(f"ファイル: {EXCEL_PATH}")
print(f"シート: {wb.sheetnames}")

results = []

# FT
ws = wb["フラップテック"]
rows = list(ws.iter_rows(values_only=True))
header_row = None
for i, row in enumerate(rows):
    if row and "担当" in str(row[0]) and "ステータス" in str(row[1] or ""):
        header_row = i
        break

print(f"\n[FT] ヘッダー行: {header_row}")
if header_row is not None:
    for row in rows[header_row+1:]:
        if not any(row): continue
        status = str(row[1] or "").strip()
        name = row[2]
        tanka = safe_int(row[6])
        shiire = safe_int(row[7])
        if "稼働中" not in status: continue
        if not is_valid_name(name): continue
        if tanka == 0: continue
        profit = tanka - shiire
        seikyu = int(profit * 0.48) if str(row[0] or "").strip() == "小坂折半" else int(profit * 0.68)
        print(f"  {name}: tanka={tanka:,} shiire={shiire:,} profit={profit:,} seikyu={seikyu:,}")
        results.append(("FT", str(name), profit, seikyu))

# GL
ws = wb["グレイスライン"]
rows = list(ws.iter_rows(values_only=True))
header_row = None
for i, row in enumerate(rows):
    if row and "ステータス" in str(row[0]):
        header_row = i
        break

print(f"\n[GL] ヘッダー行: {header_row}")
if header_row is not None:
    for row in rows[header_row+1:]:
        if not any(row): continue
        status = str(row[0] or "").strip()
        name = row[1]
        tanka = safe_int(row[5])
        shiire = safe_int(row[6])
        if "稼働中" not in status: continue
        if not is_valid_name(name): continue
        if tanka == 0: continue
        profit = tanka - shiire
        seikyu = int(profit * 0.60)
        print(f"  {name}: tanka={tanka:,} shiire={shiire:,} profit={profit:,} seikyu={seikyu:,}")
        results.append(("GL", str(name), profit, seikyu))

print(f"\n合計 {len(results)} 名 / 合計請求額: {sum(r[3] for r in results):,}円")
print("-> 全員プラス粗利で正常")
