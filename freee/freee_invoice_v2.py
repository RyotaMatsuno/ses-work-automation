"""
freee_invoice_v2.py
契約マスターExcelを正として稼働中人員の請求書をfreeeにドラフト作成。

請求ルール:
【TERRA】
  P（プロパー）: GL/FP経由以外 → 15,000円/人（税別）固定
  P（プロパー）: GL/FP経由稼働 → 請求なし
  BP: 粗利×80%
  TERRA折半: 粗利×50%
  岡本折半: 粗利×80%
【フラップテック】
  通常: 粗利×68%
  小坂折半: 粗利×48%
  岡本折半: 粗利×68%
  岡本: 粗利×68%全額払出
【グレイスライン】
  粗利×60%
"""

import os, sys, requests, json, argparse
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl

# token_managerを参照（自動リフレッシュ付き）
sys.path.insert(0, r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\freee_auth")
from token_manager import get_headers

# ===== 設定 =====
EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
FREEE_BASE_ACCT = "https://api.freee.co.jp/api/1"
FREEE_BASE_INV = "https://api.freee.co.jp/iv"
FREEE_BASE = FREEE_BASE_ACCT
COMPANY_ID = 11712776
TEMPLATE_ID = 3323260
# 源泉徴収ルール（確定）: TERRA=あり / GL=なし / FT=なし
WITHHOLDING = False
LINE_UNIT = "式"  # freee invoice API requires unit length >= 1 (empty not allowed)

def freee_headers():
    h = get_headers()
    h["Content-Type"] = "application/json"
    return h

def safe_int(v):
    """値を安全にintに変換。文字列や日付型はスキップ(0返却)"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    # 文字列・日付型はスキップ
    return 0

def is_valid_name(v):
    """氏名として有効かチェック。日付型・数値・空は除外"""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return False
    import datetime
    if isinstance(v, (datetime.datetime, datetime.date)):
        return False
    s = str(v).strip()
    if not s or s in ("NaN", "稼働中合計"):
        return False
    # 数字だけの文字列も除外
    if s.replace("/", "").replace("-", "").isdigit():
        return False
    return True

# ===== Excel読み込み =====
# [DEPRECATED 2026-06-08] Excel読込。現在は sheets_reader.load_active_entries() を使用。
def load_active_entries():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    entries = []

    # --- TERRA ---
    # ヘッダー: 担当(0) 区分(1) ステータス(2) 氏名(3) ... 案件/上位会社(6) 単価(案件)(7) ... 仕入単価(12)
    ws = wb["TERRA"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "担当" in str(row[0]):
            header_row = i
            break
    if header_row is not None:
        for row in rows[header_row+1:]:
            if not any(row): continue
            tantou   = str(row[0] or "").strip()
            kubun    = str(row[1] or "").strip()
            status   = str(row[2] or "").strip()
            name     = row[3]
            case     = str(row[6] or "").strip()
            tanka    = safe_int(row[7])
            shiire   = safe_int(row[12])

            if "稼働中" not in status: continue
            if not is_valid_name(name): continue
            name = str(name).strip()

            is_gl_ft = any(k in case for k in ["グレイスライン", "フラップテック", "GL", "FT"])
            profit = tanka - shiire

            if kubun == "P":
                if is_gl_ft:
                    continue  # 請求なし
                seikyu = 15000
                rule   = "プロパー→15,000円固定"
            elif kubun == "BP":
                if tantou == "TERRA折半":
                    seikyu = int(profit * 0.50)
                    rule   = "TERRA折半→粗利×50%"
                elif tantou == "岡本折半":
                    seikyu = int(profit * 0.80)
                    rule   = "岡本折半→粗利×80%（岡本払出あり）"
                else:
                    seikyu = int(profit * 0.80)
                    rule   = "BP→粗利×80%"
            else:
                seikyu = 15000
                rule   = "不明→15,000円固定"

            if seikyu <= 0: continue

            entries.append({
                "partner": "株式会社TERRA",
                "name": name, "profit": profit, "seikyu": seikyu,
                "rule": rule, "source": "TERRA"
            })

    # --- フラップテック ---
    # ヘッダー: 担当(0) ステータス(1) 氏名(2) 参画時期(3) 期間(4) 案件/上位(5) 案件単価(6) 仕入単価(7)
    ws = wb["フラップテック"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "担当" in str(row[0]) and "ステータス" in str(row[1] or ""):
            header_row = i
            break
    if header_row is not None:
        for row in rows[header_row+1:]:
            if not any(row): continue
            tantou  = str(row[0] or "").strip()
            status  = str(row[1] or "").strip()
            name    = row[2]
            tanka   = safe_int(row[6])   # 案件単価(上位から)
            shiire  = safe_int(row[7])   # 仕入単価(下位へ)

            if "稼働中" not in status: continue
            if not is_valid_name(name): continue
            name = str(name).strip()
            if tanka == 0: continue  # 単価未入力はスキップ

            profit = tanka - shiire

            if profit <= 0:
                print(f"  [SKIP] {name}: 粗利{profit:,}円（単価={tanka:,} 仕入={shiire:,}）")
                continue

            if tantou == "小坂折半":
                seikyu = int(profit * 0.48)
                rule   = "小坂折半→粗利×48%"
            elif tantou in ("岡本折半", "岡本"):
                seikyu = int(profit * 0.68)
                rule   = f"{tantou}→粗利×68%（岡本払出あり）"
            else:
                seikyu = int(profit * 0.68)
                rule   = "通常→粗利×68%"

            if seikyu <= 0: continue

            entries.append({
                "partner": "株式会社フラップテック",
                "name": name, "profit": profit, "seikyu": seikyu,
                "rule": rule, "source": "FT"
            })

    # --- グレイスライン ---
    # ヘッダー: ステータス(0) 氏名(1) 参画時期(2) 期間(3) 案件/上位(4) 案件単価(5) 仕入単価(6)
    ws = wb["グレイスライン"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "ステータス" in str(row[0]):
            header_row = i
            break
    if header_row is not None:
        for row in rows[header_row+1:]:
            if not any(row): continue
            status  = str(row[0] or "").strip()
            name    = row[1]
            tanka   = safe_int(row[5])   # 案件単価(上位から)
            shiire  = safe_int(row[6])   # 仕入単価(下位へ)

            if "稼働中" not in status: continue
            if not is_valid_name(name): continue
            name = str(name).strip()
            if tanka == 0: continue

            profit = tanka - shiire

            if profit <= 0:
                print(f"  [SKIP] {name}: 粗利{profit:,}円（単価={tanka:,} 仕入={shiire:,}）")
                continue

            seikyu = int(profit * 0.60)
            if seikyu <= 0: continue

            entries.append({
                "partner": "グレイスライン株式会社",
                "name": name, "profit": profit, "seikyu": seikyu,
                "rule": "GL→粗利×60%", "source": "GL"
            })

    return entries

# ===== freee: 取引先取得/作成 =====
def get_or_create_partner(name, dry_run=False):
    res = requests.get(f"{FREEE_BASE}/partners",
        headers=freee_headers(),
        params={"company_id": COMPANY_ID, "keyword": name})
    partners = res.json().get("partners", [])
    if partners: return partners[0]["id"]
    if dry_run:
        print(f"  [DRY] partner-missing:{name}")
        return 0
    res2 = requests.post(f"{FREEE_BASE}/partners",
        headers=freee_headers(),
        json={"company_id": COMPANY_ID, "name": name, "partner_type": "customer"})
    return res2.json()["partner"]["id"]

# ===== freee: 請求書ドラフト作成 =====
def fetch_existing_subjects(issue_date):
    """対象請求月(billing_date==issue_date)の既存請求書 subject 集合を返す。取得失敗時は None。"""
    target = issue_date.strftime("%Y-%m-%d")
    subs = set()
    offset = 0
    while True:
        try:
            r = requests.get(f"{FREEE_BASE_INV}/invoices",
                             headers=freee_headers(),
                             params={"company_id": COMPANY_ID, "limit": 100, "offset": offset})
        except Exception as e:
            print(f"[dedup] 一覧取得エラー: {e}")
            return None
        if r.status_code != 200:
            print(f"[dedup] 一覧取得失敗 status={r.status_code}: {r.text[:150]}")
            return None
        invs = r.json().get("invoices") or []
        for inv in invs:
            if inv.get("billing_date") == target:
                s = inv.get("subject")
                if s:
                    subs.add(s)
        if len(invs) < 100:
            break
        offset += 100
    return subs

def create_invoice(entry, issue_date, due_date, dry_run=False, existing_subjects=None):
    mon = f"{issue_date.year}年{issue_date.month}月"
    subject = f"{mon}分 業務委託料（{entry['name']}様）"
    if existing_subjects and subject in existing_subjects:
        print(f"  SKIP {entry['name']} / {entry['partner']} / 既に{mon}分の請求書あり")
        return "skip"
    partner_id = get_or_create_partner(entry["partner"], dry_run=dry_run)
    payload = {
        "company_id": COMPANY_ID,
        "partner_id":  partner_id,
        "template_id": TEMPLATE_ID,
        "billing_date": issue_date.strftime("%Y-%m-%d"),
        "payment_date": due_date.strftime("%Y-%m-%d"),
        "subject": subject,
        "payment_type": "transfer",
        "tax_entry_method": "out",
        "tax_fraction": "omit",
        "withholding_tax_entry_method": "out",
        "partner_title": "御中",
        "sending_status": "unsent",
        "lines": [{
            "type": "item",
            "description": f"業務委託料（{entry['name']}様）{mon}分",
            "quantity": 1,
            "unit": LINE_UNIT,
            "unit_price": str(entry["seikyu"]),
            "tax_rate": 10,
            "reduced_tax_rate": False,
            "withholding": WITHHOLDING
        }]
    }
    if dry_run:
        print(f"  [DRY] payload {entry['name']} / {entry['partner']}")
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return True

    res = requests.post(f"{FREEE_BASE_INV}/invoices", headers=freee_headers(), json=payload)
    if res.status_code in (200, 201):
        inv_id = res.json()["invoice"]["id"]
        print(f"  OK {entry['name']} / {entry['partner']} / {entry['seikyu']:,}円 [{entry['rule']}] -> ID:{inv_id}")
        return True
    else:
        print(f"  NG {entry['name']} / {res.status_code}: {res.text[:200]}")
        return False

# ===== メイン =====
def run(target_month=None, dry_run=False, limit=None):
    today = date.today()
    if target_month is None:
        target_month = (today.replace(day=1) + relativedelta(months=1))
    issue_date = target_month.replace(day=1)
    due_date   = issue_date + relativedelta(months=1) - relativedelta(days=1)

    print(f"=== freee請求書自動生成 v2 ===")
    if dry_run:
        print("DRY-RUN: 請求書・取引先の作成/更新は行いません")
    print(f"請求対象月: {target_month.year}年{target_month.month}月分")
    print(f"請求日: {issue_date}  支払期限: {due_date}")
    print()

    # 契約マスター = Googleスプレッドシートのみ（Excel廃止 2026-06-08）
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _root not in sys.path:
        sys.path.insert(0, _root)
    import sheets_reader
    entries = sheets_reader.load_active_entries()
    if limit is not None:
        entries = entries[:limit]
    print(f"対象人員: {len(entries)}名")
    for e in entries:
        print(f"  {e['source']} | {e['name']} | 粗利{e['profit']:,}円 | 請求{e['seikyu']:,}円 | {e['rule']}")
    print()

    existing = fetch_existing_subjects(issue_date)
    if existing is None:
        print("[dedup] 冪等チェックに失敗したため、二重請求防止のため処理を中止します。")
        return
    print(f"[dedup] {issue_date.strftime('%Y-%m-%d')} の既存請求書 {len(existing)} 件")

    ok = ng = skipped = 0
    for e in entries:
        r = create_invoice(e, issue_date, due_date, dry_run=dry_run, existing_subjects=existing)
        if r == "skip":
            skipped += 1
        elif r:
            ok += 1
        else:
            ng += 1

    print()
    if dry_run:
        print(f"=== DRY-RUN完了: 作成予定{ok}件 / SKIP{skipped}件 / エラー{ng}件 ===")
    else:
        print(f"=== 完了: 作成{ok}件 / SKIP{skipped}件 / エラー{ng}件 ===")
    print(f"-> https://secure.freee.co.jp/invoices")
    # ===== 請求書作成完了後: 契約マスターのステータスを自動更新 =====
    if ok > 0 and not dry_run:
        try:
            import sys as _sys2
            import os as _os2
            _sys2.path.insert(0, _os2.path.dirname(__file__))
            from auto_status_update import update_status_after_invoice
            invoiced_names = [e["name"] for e in entries]
            print(f"\n[auto_status] 請求書作成済み人員のステータスを稼働中に更新...")
            update_status_after_invoice(names=invoiced_names)
        except Exception as _e:
            print(f"[auto_status] ステータス更新スキップ（エラー: {_e}）")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="freee請求書自動生成 v2")
    parser.add_argument("target_month", nargs="?", help="請求対象月 YYYY-MM")
    parser.add_argument("--dry-run", action="store_true", help="payloadを表示し、POST/作成は行わない")
    parser.add_argument("--limit", type=int, help="先頭N件のみ処理")
    args = parser.parse_args()
    if args.target_month:
        y, m = map(int, args.target_month.split("-"))
        run(date(y, m, 1), dry_run=args.dry_run, limit=args.limit)
    else:
        run(dry_run=args.dry_run, limit=args.limit)
