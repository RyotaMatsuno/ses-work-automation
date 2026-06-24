# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")

import os
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
BAK_PATH = XLSX_PATH.replace(".xlsx", f"_bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
shutil.copy2(XLSX_PATH, BAK_PATH)
print(f"バックアップ: {os.path.basename(BAK_PATH)}", flush=True)

wb = openpyxl.load_workbook(XLSX_PATH)
print("既存シート:", wb.sheetnames, flush=True)

# ========== 1. 直契約シート新規作成 ==========
if "直契約" not in wb.sheetnames:
    ws_dc = wb.create_sheet("直契約")
    ws_dc["A1"] = "直契約 管理（2026-06-01作成）"
    ws_dc["A1"].font = Font(bold=True, size=12)
    ws_dc["A2"] = "請求ルール: 固定月額（粗利計算なし）　岡本支払い: 固定月額"
    headers = [
        "ステータス",
        "取引先名",
        "請求額(円/月)",
        "支払サイト",
        "岡本払出(円/月)",
        "松野実入り(円/月)",
        "開始月",
        "備考",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws_dc.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="D9D9D9")
    ws_dc["A4"] = "稼働中"
    ws_dc["B4"] = "鶴見有職研究所"
    ws_dc["C4"] = 100000
    ws_dc["D4"] = "月末締め翌月最終営業日払い"
    ws_dc["E4"] = 50000
    ws_dc["F4"] = "=C4-E4"
    ws_dc["G4"] = "2026年6月〜"
    ws_dc["H4"] = "請求先詳細は別途連携予定。岡本紹介案件（新営業スキーム適用）"
    ws_dc.column_dimensions["B"].width = 20
    ws_dc.column_dimensions["D"].width = 28
    ws_dc.column_dimensions["H"].width = 45
    print("[OK] 直契約シート作成", flush=True)
else:
    print("[SKIP] 直契約シートは既存", flush=True)

# ========== 2. 月次サマリーに実績追記 ==========
ws_sum = wb["月次サマリー"]
last_row = ws_sum.max_row

ws_sum.cell(row=last_row + 2, column=1, value="■ 2026年 月次実入り実績（確定値）").font = Font(bold=True)
headers2 = ["対象月", "入金月", "総実入り(円)", "岡本払出(円)", "木原分(円)", "松野手残り(円)", "備考"]
for i, h in enumerate(headers2, 1):
    cell = ws_sum.cell(row=last_row + 3, column=i, value=h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", start_color="D9D9D9")

row6 = last_row + 4
ws_sum.cell(row=row6, column=1, value="4月稼働分")
ws_sum.cell(row=row6, column=2, value="2026年6月")
ws_sum.cell(row=row6, column=3, value=787202)
ws_sum.cell(row=row6, column=4, value=70800)
ws_sum.cell(row=row6, column=5, value=116000)
ws_sum.cell(row=row6, column=6, value=f"=C{row6}-D{row6}-E{row6}")
ws_sum.cell(row=row6, column=7, value="岡本払出70,800・木原116,000控除後")

row7 = last_row + 5
ws_sum.cell(row=row7, column=1, value="5月稼働分")
ws_sum.cell(row=row7, column=2, value="2026年7月")
ws_sum.cell(row=row7, column=3, value=825282)
ws_sum.cell(row=row7, column=4, value=80320)
ws_sum.cell(row=row7, column=5, value=116000)
ws_sum.cell(row=row7, column=6, value=f"=C{row7}-D{row7}-E{row7}")
ws_sum.cell(row=row7, column=7, value="原昌志5月末終了→7月以降木原分96,000に減少の可能性あり（要確認）")
print(f"[OK] 月次サマリーに6・7月実績追記（行{row6},{row7}）", flush=True)

# ========== 3. FTシートに新営業報酬ルール追記 ==========
ws_ft = wb["フラップテック"]
current_rule = ws_ft["A2"].value or ""
new_note = (
    " / 【新営業報酬ルール(2026-05確定)】"
    "紹介のみ=請求額x20% / 自己完結=請求額x90% / "
    "松野と成約=松野50%・岡本10%・新営業40% / "
    "岡本と成約=岡本60%・新営業40% / "
    "小坂代表と成約=粗利x50%請求→新営業に請求額x90%払出"
)
ws_ft["A2"] = current_rule + new_note
print("[OK] FT新営業報酬ルール追記", flush=True)

# ========== 4. 備考・メモシート ==========
memo_name = "備考・メモ"
if memo_name not in wb.sheetnames:
    ws_memo = wb.create_sheet(memo_name)
else:
    ws_memo = wb[memo_name]

rows_memo = [
    ("■ 法人化タイミング（2026-06-01確定）", True),
    ("推奨タイミング: 2027年上半期（1〜3月）", False),
    ("今年中のアクション: 専従者給与届出のみ先行（年末年始前に提出）", False),
    (
        "理由: 2026年内法人化との差額は約32〜47万円。FT直接契約移管（増収+47〜63万）と合わせて2027年1月設立が最適解",
        False,
    ),
    ("", False),
    ("■ FP小坂さん 所属共有スキーム（2026-05確定）", True),
    ("内容: 互いの1社下案件の所属情報を共有→直接調整することで1社下の抜き分を分配", False),
    (
        "パターンA（小坂案件）: 松野岡本が所属と直接交渉→1社下抜き分×80%が松野岡本・20%が小坂（FP件数カウントなし）",
        False,
    ),
    (
        "パターンB（松野岡本案件）: 小坂が所属情報提供→元々の利益そのまま+1社下抜き分×20%が小坂（FP件数カウントあり）",
        False,
    ),
    ("", False),
    ("■ 直契約（鶴見有職研究所）（2026-06-01確定）", True),
    ("月10万固定請求・翌月最終営業日払い・岡本払出5万/月・2026年6月〜", False),
    ("請求先詳細は別途連携予定", False),
]
for i, (text, bold) in enumerate(rows_memo, 1):
    cell = ws_memo.cell(row=i, column=1, value=text)
    if bold:
        cell.font = Font(bold=True)
ws_memo.column_dimensions["A"].width = 90
print("[OK] 備考・メモシート更新", flush=True)

wb.save(XLSX_PATH)
print(f"\n[DONE] 保存完了: {XLSX_PATH}", flush=True)
print("最終シート一覧:", wb.sheetnames, flush=True)
