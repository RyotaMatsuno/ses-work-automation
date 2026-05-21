
import openpyxl, shutil
from datetime import datetime

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
BAK = EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
shutil.copy2(EXCEL_PATH, BAK)

wb = openpyxl.load_workbook(EXCEL_PATH)

# --- TERRA修正 ---
ws_terra = wb["TERRA"]

# 行37(鶴川): TERRAから削除
# 行38(斎藤): 氏名を「齋藤よしまさ」に修正
# ただし行37を先に削除すると38が37になるので、後ろから処理

# 行38（斎藤 → 齋藤よしまさ）修正
ws_terra.cell(38, 4).value = "齋藤よしまさ"
print("TERRA 行38: 斎藤 → 齋藤よしまさ")

# 行37（鶴川）を削除
ws_terra.delete_rows(37, 1)
print("TERRA 行37（鶴川）削除")

# --- FTに鶴川を追加 ---
ws_ft = wb["フラップテック"]

# 現在の最終データ行確認
last_row = 13  # 吉田祥平が行13、合計行が14
ws_ft.insert_rows(14, 1)
ws_ft.cell(14, 1).value = None          # 担当
ws_ft.cell(14, 2).value = "稼働中"       # ステータス
ws_ft.cell(14, 3).value = "鶴川"        # 氏名
ws_ft.cell(14, 4).value = "2026/5"      # 参画時期
ws_ft.cell(14, 5).value = "長期"        # 期間
ws_ft.cell(14, 6).value = "（確認中）"  # 案件
print("FT 行14: 鶴川 追加（稼働中 2026/5）")

wb.save(EXCEL_PATH)
print(f"保存完了 / バックアップ: {BAK.split('\\')[-1]}")
