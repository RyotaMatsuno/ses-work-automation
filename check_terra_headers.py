# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)

# ============================================================
# TERRAシート 備考・実入り情報を全員分入力
# ============================================================
ws_t = wb["TERRA"]

terra_row_map = {}
for i, row in enumerate(ws_t.iter_rows(min_row=4, values_only=True), start=4):
    if row[3]:
        terra_row_map[str(row[3])] = i

# 実入り列はcol17（R列）、備考はcol18（S列）
# 岡本払出はcol17(Q)、実入りはcol18(R)…ヘッダーを確認
headers = [ws_t.cell(row=3, column=i).value for i in range(1, 22)]
print("TERRAヘッダー:", headers, flush=True)
