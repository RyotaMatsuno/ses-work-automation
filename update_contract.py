# -*- coding: utf-8 -*-
# 契約マスターに以下を追加・更新:
# 1. 「直契約」シート新規作成（鶴見有職研究所）
# 2. 月次予測シートに6月・7月実入りサマリーを追記
# 3. FTシートの請求ルール欄に新営業報酬ルールを追記
# 4. メモシートがなければ「備考・メモ」シートを新規作成して法人化・FP小坂スキームを記録

import os
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
BAK_PATH = XLSX_PATH.replace(".xlsx", f"_bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

# バックアップ
shutil.copy2(XLSX_PATH, BAK_PATH)
print(f"バックアップ: {os.path.basename(BAK_PATH)}", flush=True)

wb = openpyxl.load_workbook(XLSX_PATH)
print("既存シート:", wb.sheetnames, flush=True)

# ========== 1. 「直契約」シート新規作成 ==========
if "直契約" not in wb.sheetnames:
    ws_dc = wb.create_sheet("直契約")
    # ヘッダー
    ws_dc["A1"] = "直契約 管理（2026-06-01作成）"
    ws_dc["A1"].font = Font(bold=True, size=12)
    ws_dc["A2"] = "請求ルール: 固定月額（粗利計算なし・そのまま請求）　岡本支払い: 固定月額"
    ws_dc["A3"] = "ステータス"
    ws_dc["B3"] = "取引先名"
    ws_dc["C3"] = "請求額(円/月)"
    ws_dc["D3"] = "支払サイト"
    ws_dc["E3"] = "岡本払出(円/月)"
    ws_dc["F3"] = "松野実入り(円/月)"
    ws_dc["G3"] = "開始月"
    ws_dc["H3"] = "備考"
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws_dc[f"{col}3"].font = Font(bold=True)
        ws_dc[f"{col}3"].fill = PatternFill("solid", start_color="D9D9D9")

    # 鶴見有職研究所
    ws_dc["A4"] = "稼働中"
    ws_dc["B4"] = "鶴見有職研究所"
    ws_dc["C4"] = 100000
    ws_dc["D4"] = "月末締め翌月最終営業日払い"
    ws_dc["E4"] = 50000
    ws_dc["F4"] = "=C4-E4"
    ws_dc["G4"] = "2026年6月〜"
    ws_dc["H4"] = "請求先詳細は別途連携予定。岡本紹介案件（新営業スキーム適用）"

    ws_dc.column_dimensions["B"].width = 20
    ws_dc.column_dimensions["D"].width = 25
    ws_dc.column_dimensions["H"].width = 40
    print("✅ 直契約シート作成", flush=True)
else:
    print("直契約シートは既存", flush=True)

# ========== 2. 月次サマリーシートに実績行を追記 ==========
ws_sum = wb["月次サマリー"]
last_row = ws_sum.max_row

# 実績サマリーを末尾に追加
ws_sum.cell(row=last_row + 2, column=1, value="■ 2026年 月次実入り実績（確定値）")
ws_sum.cell(row=last_row + 2, column=1).font = Font(bold=True)

headers = ["対象月", "入金月", "総実入り(円)", "岡本払出(円)", "木原分(円)", "松野手残り(円)", "備考"]
for i, h in enumerate(headers, 1):
    cell = ws_sum.cell(row=last_row + 3, column=i, value=h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", start_color="D9D9D9")

# 6月入金（4月稼働分）
row6 = last_row + 4
ws_sum.cell(row=row6, column=1, value="4月稼働分")
ws_sum.cell(row=row6, column=2, value="2026年6月")
ws_sum.cell(row=row6, column=3, value=787202)
ws_sum.cell(row=row6, column=4, value=70800)
ws_sum.cell(row=row6, column=5, value=116000)
ws_sum.cell(row=row6, column=6, value="=C{r}-D{r}-E{r}".format(r=row6))
ws_sum.cell(row=row6, column=7, value="岡本払出70,800・木原116,000控除後")

# 7月入金（5月稼働分）
row7 = last_row + 5
ws_sum.cell(row=row7, column=1, value="5月稼働分")
ws_sum.cell(row=row7, column=2, value="2026年7月")
ws_sum.cell(row=row7, column=3, value=825282)
ws_sum.cell(row=row7, column=4, value=80320)
ws_sum.cell(row=row7, column=5, value=116000)
ws_sum.cell(row=row7, column=6, value="=C{r}-D{r}-E{r}".format(r=row7))
ws_sum.cell(row=row7, column=7, value="原昌志5月末終了→7月以降木原分が96,000に減少の可能性あり（要確認）")
print(f"✅ 月次サマリーに6月・7月実績追記（行{row6},{row7}）", flush=True)

# ========== 3. FTシートの請求ルール欄に新営業報酬ルールを追記 ==========
ws_ft = wb["フラップテック"]
# A2が請求ルール行
current_rule = ws_ft["A2"].value or ""
new_rule_note = (
    "\n【新営業報酬ルール（2026-05確定）】"
    "①紹介のみ=請求額×20%　"
    "②自己完結=請求額×90%　"
    "③松野と成約=松野50%/岡本10%/新営業40%　"
    "④岡本と成約=岡本60%/新営業40%　"
    "⑤小坂代表と成約=粗利×50%請求→新営業に請求額×90%払出"
)
ws_ft["A2"] = current_rule + new_rule_note
print("✅ FT新営業報酬ルール追記", flush=True)

# ========== 4. 備考・メモシート作成 ==========
memo_name = "備考・メモ"
if memo_name not in wb.sheetnames:
    ws_memo = wb.create_sheet(memo_name)
else:
    ws_memo = wb[memo_name]

# 法人化タイミング
ws_memo["A1"] = "■ 法人化タイミング（2026-06-01確定）"
ws_memo["A1"].font = Font(bold=True)
ws_memo["A2"] = "推奨タイミング: 2027年上半期（1〜3月）"
ws_memo["A3"] = "今年中のアクション: 専従者給与届出のみ先行（年末年始前に提出）"
ws_memo["A4"] = (
    "理由: 2026年内法人化との差額は約32〜47万円。FT直接契約移管（増収+47〜63万）と合わせて2027年1月設立が最適解"
)
ws_memo["A5"] = "2026年内法人化スケジュール案: 7月準備→8〜9月設立→9〜10月FT移管→12月末専従者届出不要（役員報酬に移行）"

# FP小坂スキーム
ws_memo["A7"] = "■ FP小坂さん 所属共有スキーム（2026-05確定）"
ws_memo["A7"].font = Font(bold=True)
ws_memo["A8"] = "内容: お互いの1社下案件の所属情報を共有し合い、直接調整することで1社下の抜き分を分配"
ws_memo["A9"] = (
    "パターンA（小坂案件）: 松野岡本が所属と直接交渉 → 1社下の抜き分×80%が松野岡本・20%が小坂（件数カウントなし）"
)
ws_memo["A10"] = (
    "パターンB（松野岡本案件）: 小坂が所属情報提供 → 元々の利益はそのまま+1社下抜き分×20%が小坂（件数カウントあり）"
)
ws_memo["A11"] = "精算: 成約時都度。1社下の抜き幅は申告ベース"

ws_memo.column_dimensions["A"].width = 80
print("✅ 備考・メモシート更新", flush=True)

# 保存
wb.save(XLSX_PATH)
print(f"\n✅ 保存完了: {XLSX_PATH}", flush=True)
print("シート一覧:", wb.sheetnames, flush=True)
