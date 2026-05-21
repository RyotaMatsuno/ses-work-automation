
"""
Googleスプレッドシートに契約マスターExcelデータを書き込む
"""
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from googleapiclient.discovery import build

CREDS_PATH  = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\google_credentials.json"
EXCEL_PATH  = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
SS_ID = "1ORBtxtGqLAwv3YU8CGeLX7gWFgvKOivMTCZZiWtYGfI"
SCOPES = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

creds = Credentials.from_service_account_file(CREDS_PATH, scopes=SCOPES)
gc = gspread.authorize(creds)

# Service AccountにスプレッドシートのEditor権限を付与（Drive API経由）
drive_svc = build("drive", "v3", credentials=creds)
drive_svc.permissions().create(
    fileId=SS_ID,
    body={"type": "user", "role": "writer", "emailAddress": "ses-965@ses-work-automation.iam.gserviceaccount.com"},
    sendNotificationEmail=False
).execute()
print("Service Account に editor 権限付与")

# スプレッドシートを開く
ss = gc.open_by_key(SS_ID)
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

existing = [ws.title for ws in ss.worksheets()]

for i, sheet_name in enumerate(wb.sheetnames):
    ws_excel = wb[sheet_name]
    rows = list(ws_excel.iter_rows(values_only=True))

    if sheet_name in existing:
        gs = ss.worksheet(sheet_name)
        gs.clear()
    elif i == 0 and "Sheet1" in existing:
        gs = ss.worksheet("Sheet1")
        gs.update_title(sheet_name)
    else:
        gs = ss.add_worksheet(title=sheet_name, rows=max(len(rows)+20, 50), cols=25)

    data = []
    for row in rows:
        data.append([str(v) if v is not None else "" for v in row])

    if data:
        gs.update(data, "A1")
    print(f"  [{sheet_name}] {len(rows)}行 書込完了")

# IDを保存
with open(r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\spreadsheet_id.txt", "w") as f:
    f.write(SS_ID)

print(f"\n完了! URL: https://docs.google.com/spreadsheets/d/{SS_ID}")
