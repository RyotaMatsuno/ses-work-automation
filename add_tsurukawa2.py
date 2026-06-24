# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_ft = wb["フラップテック"]

# 結合セルを確認
print("結合セル一覧:", list(ws_ft.merged_cells.ranges)[:10], flush=True)

# 最終データ行の次（遠藤健太が16行目）→17行目に追加
new_row = 17

try:
    ws_ft.cell(row=new_row, column=1, value="岡本全額")
    ws_ft.cell(row=new_row, column=2, value="稼働中")
    ws_ft.cell(row=new_row, column=3, value="鶴川慶三")
    ws_ft.cell(row=new_row, column=4, value="2026/6")
    ws_ft.cell(row=new_row, column=5, value="長期")
    ws_ft.cell(row=new_row, column=6, value="アバンテック")
    ws_ft.cell(row=new_row, column=7, value=720000)
    ws_ft.cell(row=new_row, column=8, value=650000)
    ws_ft.cell(row=new_row, column=9, value=f"=G{new_row}-H{new_row}")
    ws_ft.cell(row=new_row, column=10, value=f"=I{new_row}*0.68")
    ws_ft.cell(row=new_row, column=11, value=f"=J{new_row}")
    ws_ft.cell(row=new_row, column=12, value=0)
    ws_ft.cell(row=new_row, column=13, value=45)
    ws_ft.cell(row=new_row, column=14, value="上位からくる")
    ws_ft.cell(row=new_row, column=16, value=0)
    ws_ft.cell(row=new_row, column=17, value="単月更新")
    ws_ft.cell(
        row=new_row,
        column=18,
        value="岡本全額払出。粗利7万×68%=47,600円を全額岡本へ払出。松野実入り0円。6月入場→8月15日初回入金",
    )
    print(f"[OK] 鶴川慶三 行{new_row}追加", flush=True)
except AttributeError as e:
    print(f"行{new_row}でエラー: {e} → 別の行を試します", flush=True)
    # 結合外の行を探す
    for try_row in range(18, 25):
        try:
            ws_ft.cell(row=try_row, column=1, value="岡本全額")
            ws_ft.cell(row=try_row, column=2, value="稼働中")
            ws_ft.cell(row=try_row, column=3, value="鶴川慶三")
            ws_ft.cell(row=try_row, column=4, value="2026/6")
            ws_ft.cell(row=try_row, column=5, value="長期")
            ws_ft.cell(row=try_row, column=6, value="アバンテック")
            ws_ft.cell(row=try_row, column=7, value=720000)
            ws_ft.cell(row=try_row, column=8, value=650000)
            ws_ft.cell(row=try_row, column=9, value=f"=G{try_row}-H{try_row}")
            ws_ft.cell(row=try_row, column=10, value=f"=I{try_row}*0.68")
            ws_ft.cell(row=try_row, column=11, value=f"=J{try_row}")
            ws_ft.cell(row=try_row, column=12, value=0)
            ws_ft.cell(row=try_row, column=13, value=45)
            ws_ft.cell(row=try_row, column=14, value="上位からくる")
            ws_ft.cell(row=try_row, column=16, value=0)
            ws_ft.cell(row=try_row, column=17, value="単月更新")
            ws_ft.cell(
                row=try_row,
                column=18,
                value="岡本全額払出。粗利7万×68%=47,600円を全額岡本へ払出。松野実入り0円。6月入場→8月15日初回入金",
            )
            print(f"[OK] 鶴川慶三 行{try_row}に追加成功", flush=True)
            break
        except AttributeError:
            print(f"  行{try_row}もマージ済みのためスキップ", flush=True)
            continue

wb.save(XLSX_PATH)
print("[DONE]", flush=True)
