import openpyxl

f = r"C:\Users\ma_py\Downloads\単価マスター_v5.xlsx"
wb = openpyxl.load_workbook(f)
print("シート一覧:", wb.sheetnames)

for sh in wb.sheetnames:
    ws = wb[sh]
    print(f"\n====== {sh} ======")
    # ヘッダー行(1-3行目)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), 1):
        print(f"行{i}: {row}")
    # データ行サンプル(4-6行目)
    print("--- データサンプル ---")
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=7, values_only=True), 4):
        print(f"行{i}: {row}")
