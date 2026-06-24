# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_t = wb["TERRA"]

for i, row in enumerate(ws_t.iter_rows(min_row=4, values_only=True), start=4):
    if row[3] == "齋藤よしまさ":
        ws_t.cell(
            row=i,
            column=18,
            value="初月5/15入場 0.61人月 TERRA請求=15,000×0.61=9,150円（7月入金）。6月稼働〜フル15,000円",
        )
        print("[OK] 齋藤よしまさ 備考更新: 0.61人月", flush=True)
        break

wb.save(XLSX_PATH)
print("[DONE]", flush=True)
