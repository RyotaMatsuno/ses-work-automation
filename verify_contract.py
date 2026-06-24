# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
print("シート一覧:", wb.sheetnames, flush=True)

# 直契約シート確認
if "直契約" in wb.sheetnames:
    ws = wb["直契約"]
    print("\n=== 直契約 ===")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)

# 月次サマリー末尾確認
ws_s = wb["月次サマリー"]
print(f"\n=== 月次サマリー（末尾5行 / 最大行:{ws_s.max_row}）===")
for row in ws_s.iter_rows(min_row=ws_s.max_row - 4, max_row=ws_s.max_row, values_only=True):
    print(row)

# 備考メモシート確認
if "備考・メモ" in wb.sheetnames:
    ws_m = wb["備考・メモ"]
    print("\n=== 備考・メモ ===")
    for row in ws_m.iter_rows(min_row=1, max_row=13, values_only=True):
        if any(v for v in row):
            print(row[0])
