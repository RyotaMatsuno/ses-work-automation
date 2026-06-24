# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

print("=== TERRA シート ===", flush=True)
ws = wb["TERRA"]
headers = [ws.cell(row=3, column=i).value for i in range(1, 22)]
print("ヘッダー:", headers[:18], flush=True)
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[2] and row[2] in ["稼働中", "入場前"]:
        print(row[:18], flush=True)

print("\n=== GL シート ===", flush=True)
ws = wb["グレイスライン"]
headers = [ws.cell(row=3, column=i).value for i in range(1, 18)]
print("ヘッダー:", headers, flush=True)
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] and "稼働中" in str(row[0]):
        print(row[:15], flush=True)

print("\n=== FT シート ===", flush=True)
ws = wb["フラップテック"]
headers = [ws.cell(row=3, column=i).value for i in range(1, 20)]
print("ヘッダー:", headers, flush=True)
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[1] and "稼働中" in str(row[1]):
        print(row[:18], flush=True)

print("\n=== 直契約 シート ===", flush=True)
ws = wb["直契約"]
for row in ws.iter_rows(min_row=3, values_only=True):
    if any(v for v in row):
        print(row, flush=True)
