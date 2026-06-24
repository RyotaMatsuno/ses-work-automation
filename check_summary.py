# -*- coding: utf-8 -*-
# 月次サマリーと月次予測の現状を確認
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)

for sname in ["月次サマリー", "月次予測"]:
    ws = wb[sname]
    print(f"\n=== {sname} (最大行:{ws.max_row} 最大列:{ws.max_column}) ===")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            print(row[:8])
