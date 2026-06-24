# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import shutil
from datetime import datetime

import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
BAK = XLSX_PATH.replace(".xlsx", f"_bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
shutil.copy2(XLSX_PATH, BAK)
print(f"バックアップ: {BAK}", flush=True)

wb = openpyxl.load_workbook(XLSX_PATH)

# ============================================================
# 1. TERRAシート 修正
# ============================================================
ws_t = wb["TERRA"]

# 行番号マップを作成（氏名→行）
terra_row_map = {}
for i, row in enumerate(ws_t.iter_rows(min_row=4, values_only=True), start=4):
    if row[3]:
        terra_row_map[str(row[3])] = i

print(f"TERRAマップ: {list(terra_row_map.keys())}", flush=True)


def t_set(name, col, val):
    r = terra_row_map.get(name)
    if r:
        ws_t.cell(row=r, column=col, value=val)
        print(f"  TERRA[{name}] col{col} = {val}", flush=True)
    else:
        print(f"  TERRA[{name}] NOT FOUND", flush=True)


# 大本（5月末終了→ステータス更新、期間修正）
t_set("大本", 3, "5月末終了")
# 魚谷（4月終了）- ステータス確認が必要なので備考追加のみ
# ※魚谷は「4月終了」と書いてあるが稼働中 → 退場済みに変更
t_set("魚谷", 3, "退場済み")

# 吉田祥平（TERRA側）入場前→稼働中に更新
t_set("吉田祥平", 3, "稼働中")

# 川崎(新) 単価・サイト未入力 → 備考欄に要確認メモ
t_set("川崎(新)", 18, "単価・サイト未確認・要更新")

# 橋詰(新) サイト未入力 → 備考欄
t_set("橋詰(新)", 18, "TERRA請求額・担当者要確認")

# 佐々木(入) TERRA請求額空欄 → col16に設定
# TERRAプロパーなら15,000
t_set("佐々木(入)", 16, 15000)
t_set("佐々木(入)", 18, "TERRAプロパー扱いで15,000を仮入力。確認必要")

# 齋藤よしまさ 単価・全項目空欄
t_set("齋藤よしまさ", 18, "全情報未入力・要確認")

print("[OK] TERRAシート更新", flush=True)

# ============================================================
# 2. GLシート 粗利・GL請求額・実入り・サイト 数式入力
# ============================================================
ws_gl = wb["グレイスライン"]

# GLの行マップ
gl_row_map = {}
for i, row in enumerate(ws_gl.iter_rows(min_row=4, values_only=True), start=4):
    if row[1]:
        gl_row_map[str(row[1])] = i

print(f"GLマップ: {list(gl_row_map.keys())}", flush=True)

# ヘッダー確認: col1=ステータス,2=氏名,3=参画,4=期間,5=案件,6=案件単価,7=仕入単価,8=粗利,9=GL請求,10=実入り,11=サイト,...,15=木原
# 石崎春光 サイト30, 木原0
# 山内清   サイト45, 木原10000
# 荒井大輝 サイト45, 木原10000

gl_data = {
    "石崎春光": {"site": 30, "kihara": 0},
    "山内清": {"site": 45, "kihara": 10000},
    "荒井大輝": {"site": 45, "kihara": 10000},
}

for name, d in gl_data.items():
    r = gl_row_map.get(name)
    if not r:
        print(f"  GL[{name}] NOT FOUND", flush=True)
        continue
    # 粗利 = 案件単価(col6) - 仕入単価(col7)
    ws_gl.cell(row=r, column=8, value=f"=F{r}-G{r}")
    # GL請求額 = 粗利×60%
    ws_gl.cell(row=r, column=9, value=f"=H{r}*0.6")
    # 実入り = GL請求 - 木原
    ws_gl.cell(row=r, column=10, value=f"=I{r}-{d['kihara']}")
    # 支払サイト
    ws_gl.cell(row=r, column=11, value=d["site"])
    print(f"  GL[{name}] 粗利・GL請求・実入り・サイト{d['site']}日 入力", flush=True)

print("[OK] GLシート更新", flush=True)

# ============================================================
# 3. FTシート 粗利・FT請求額・岡本払出・実入り 数式入力
# ============================================================
ws_ft = wb["フラップテック"]

# FTヘッダー: 1=担当,2=ステータス,3=氏名,4=参画,5=期間,6=案件,7=案件単価,8=仕入単価,9=粗利,10=FT請求,11=岡本払出,12=実入り,13=サイト,...,16=木原
# 区分: 通常=粗利×68%, 小坂折半=粗利×48%, 岡本折半=FT請求÷2が岡本払出

ft_row_map = {}
for i, row in enumerate(ws_ft.iter_rows(min_row=4, values_only=True), start=4):
    if row[2]:
        ft_row_map[str(row[2])] = i

print(f"FTマップ: {list(ft_row_map.keys())}", flush=True)

ft_configs = {
    "笠井健太": {"tanto": "通常", "rate": 0.68},
    "木村勇太": {"tanto": "通常", "rate": 0.68},
    "加藤": {"tanto": "小坂折半", "rate": 0.48},
    "川崎健太": {"tanto": "通常", "rate": 0.68},
    "田中みさ": {"tanto": "通常", "rate": 0.68},
    "立野和紀": {"tanto": "通常", "rate": 0.68},
    "佐々木駿": {"tanto": "通常", "rate": 0.68},
    "橋本奈緒": {"tanto": "岡本折半", "rate": 0.68},
    "吉田祥平": {"tanto": "小坂折半", "rate": 0.48},
    "遠藤健太": {"tanto": "岡本折半", "rate": 0.68},
}

for name, cfg in ft_configs.items():
    r = ft_row_map.get(name)
    if not r:
        print(f"  FT[{name}] NOT FOUND", flush=True)
        continue

    rate = cfg["rate"]
    tanto = cfg["tanto"]

    # 粗利 = 案件単価(G) - 仕入単価(H)
    ws_ft.cell(row=r, column=9, value=f"=G{r}-H{r}")

    # FT請求額
    ws_ft.cell(row=r, column=10, value=f"=I{r}*{rate}")

    # 岡本払出・実入り
    if tanto == "岡本折半":
        ws_ft.cell(row=r, column=11, value=f"=J{r}*0.5")  # 岡本払出 = FT請求÷2
        ws_ft.cell(row=r, column=12, value=f"=J{r}-K{r}-P{r}")  # 実入り = FT請求-岡本払出-木原
    elif tanto == "岡本全額":
        ws_ft.cell(row=r, column=11, value=f"=J{r}")  # 岡本払出 = FT請求全額
        ws_ft.cell(row=r, column=12, value=0)
    else:
        ws_ft.cell(row=r, column=11, value=0)
        ws_ft.cell(row=r, column=12, value=f"=J{r}-P{r}")  # 実入り = FT請求-木原

    print(f"  FT[{name}] ({tanto} ×{rate}) 数式入力", flush=True)

# 鶴川慶三（岡本全額）
r_tc = ft_row_map.get("鶴川慶三")
if r_tc:
    ws_ft.cell(row=r_tc, column=9, value=f"=G{r_tc}-H{r_tc}")
    ws_ft.cell(row=r_tc, column=10, value=f"=I{r_tc}*0.68")
    ws_ft.cell(row=r_tc, column=11, value=f"=J{r_tc}")  # 全額岡本払出
    ws_ft.cell(row=r_tc, column=12, value=0)
    print("  FT[鶴川慶三] (岡本全額) 数式入力", flush=True)

# 吉田祥平の木原分を0にセット（現在Noneのため）
r_ys = ft_row_map.get("吉田祥平")
if r_ys:
    ws_ft.cell(row=r_ys, column=16, value=0)
    print("  FT[吉田祥平] 木原分=0 設定", flush=True)

print("[OK] FTシート更新", flush=True)

# ============================================================
# 4. 直契約シート 松野実入り数式
# ============================================================
ws_dc = wb["直契約"]
for i, row in enumerate(ws_dc.iter_rows(min_row=4, values_only=True), start=4):
    if row[1] == "鶴見有職研究所":
        ws_dc.cell(row=i, column=6, value=f"=C{i}-E{i}")
        print("  直契約[鶴見] 実入り数式入力", flush=True)
        break

# ============================================================
# 保存
# ============================================================
wb.save(XLSX_PATH)
print(f"\n[DONE] 保存完了: {XLSX_PATH}", flush=True)
