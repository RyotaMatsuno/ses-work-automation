# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import io
import os

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

KEY = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\google_credentials.json"
ORIG = "1nSbYDRA9nzezksBDZacTpR4udEIhWOtixcwD7jTXLFk"
FMT = "1TyKVdrwzUdURam5v9HG4A_RhyKJUhDmX"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
creds = Credentials.from_service_account_file(KEY, scopes=SCOPES)

# 1) 元シート(ネイティブ)を全ダンプ
gc = gspread.authorize(creds)
ss = gc.open_by_key(ORIG)
ws = ss.worksheet("経験者用")
vals = ws.get_all_values()
print("=== 元スキルシート [経験者用] 全行ダンプ ===", flush=True)
print(f"(rows={len(vals)})", flush=True)
for i, row in enumerate(vals, 1):
    # 空でない行のみ、列インデックス付きで
    cells = [f"{chr(65 + j)}={c}" for j, c in enumerate(row) if str(c).strip()]
    if cells:
        print(f"R{i}: " + " | ".join(cells), flush=True)

# 2) FMT .xlsx をDrive APIでダウンロード → 構造把握
print("\n=== フォーマット .xlsx メタ + 構造 ===", flush=True)
drive = build("drive", "v3", credentials=creds)
meta = drive.files().get(fileId=FMT, fields="id,name,mimeType,size,owners(emailAddress)").execute()
print(f"name={meta.get('name')} mime={meta.get('mimeType')} size={meta.get('size')}", flush=True)
print(f"owners={[o.get('emailAddress') for o in meta.get('owners', [])]}", flush=True)

req = drive.files().get_media(fileId=FMT)
buf = io.BytesIO()
dl = MediaIoBaseDownload(buf, req)
done = False
while not done:
    _, done = dl.next_chunk()
local_path = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\ys_format_dl.xlsx"
with open(local_path, "wb") as f:
    f.write(buf.getvalue())
print(f"DL OK -> {local_path} ({os.path.getsize(local_path)} bytes)", flush=True)

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --quiet")
    import openpyxl
wb = openpyxl.load_workbook(local_path, data_only=False)
print(f"sheets={wb.sheetnames}", flush=True)
for sn in wb.sheetnames:
    wsx = wb[sn]
    print(f"\n--- [{sn}] dims={wsx.dimensions} max_row={wsx.max_row} max_col={wsx.max_column} ---", flush=True)
    for row in wsx.iter_rows():
        line = []
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                line.append(f"{cell.coordinate}={repr(cell.value)}")
        if line:
            print("  " + " | ".join(line), flush=True)
print("\n[DONE]", flush=True)
