# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_t = wb["TERRA"]

# 行マップ
terra_row_map = {}
for i, row in enumerate(ws_t.iter_rows(min_row=4, values_only=True), start=4):
    if row[3]:
        terra_row_map[str(row[3])] = i

updates = {
    "川崎(新)": {"tanka": 700000, "site": 45, "terra_req": 15000},
    "橋詰(新)": {"tanka": 450000, "site": 45, "terra_req": 15000},
    "齋藤よしまさ": {"tanka": 430000, "site": 45, "terra_req": 15000},
}

# ヘッダー確認: H=案件単価(col8), I=支払サイト(col9), P=TERRA請求額(col16), R=備考(col18)
for name, d in updates.items():
    r = terra_row_map.get(name)
    if not r:
        print(f"NOT FOUND: {name}", flush=True)
        continue
    ws_t.cell(row=r, column=8, value=d["tanka"])
    ws_t.cell(row=r, column=9, value=d["site"])
    ws_t.cell(row=r, column=16, value=d["terra_req"])
    ws_t.cell(row=r, column=18, value=None)  # 要確認メモを削除
    print(f"[OK] {name}: 単価{d['tanka']:,}円 サイト{d['site']}日 TERRA請求{d['terra_req']:,}円", flush=True)

wb.save(XLSX_PATH)
print("\n[DONE] 保存完了", flush=True)
