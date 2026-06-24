import json

import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx")

result = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append({"row": i, "data": list(row)})
    result[sheet_name] = rows

with open(r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\excel_dump.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, default=str, indent=2)

print("Done")
