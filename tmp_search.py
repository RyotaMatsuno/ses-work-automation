
import openpyxl, json

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# 全シートで鶴川・斎藤を検索
targets = ["鶴川", "斎藤"]
for sname in wb.sheetnames:
    ws = wb[sname]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell and any(t in str(cell) for t in targets):
                print(f"HIT: シート={sname} 行{i} 値={cell} | 行全体={row[:6]}")
print("検索完了")
