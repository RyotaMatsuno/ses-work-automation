
import openpyxl, json

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# 全シートで斎藤・鶴川を検索
keywords = ["斎藤", "齋藤", "鶴川"]
for sname in wb.sheetnames:
    ws = wb[sname]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for j, cell in enumerate(row):
            if cell and any(k in str(cell) for k in keywords):
                print(f"HIT: sheet={sname} 行{i} 列{j+1} 値={cell} | 行全体={[str(v)[:20] if v else None for v in row[:8]]}")
