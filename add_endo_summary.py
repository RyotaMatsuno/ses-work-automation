# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)

# 月次サマリーの入金予測行に遠藤健太分を注記追加
ws_s = wb["月次サマリー"]
last_row = ws_s.max_row
# 9月入金行を追記
new_row = last_row + 1
ws_s.cell(row=new_row, column=1, value="7月稼働分（遠藤健太含む）")
ws_s.cell(row=new_row, column=2, value="2026年9月15日")
ws_s.cell(row=new_row, column=3, value="遠藤健太FT分: 粗利5万×68%=34,000→岡本折半後松野実入り17,000円を含む")
ws_s.cell(row=new_row, column=7, value="遠藤健太（FT岡本折半）初回入金。9/15確定")

wb.save(XLSX_PATH)
print("[DONE] 月次サマリーに9月入金メモ追記", flush=True)
