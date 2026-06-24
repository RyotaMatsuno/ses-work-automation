# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# 入金予測シートの中身を確認
ws = wb["入金予測"]
print(f"=== 入金予測シート (最大行:{ws.max_row}) ===", flush=True)
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    if any(v is not None for v in row):
        print(row[:10], flush=True)

print("\n--- 2026年度行を探す ---", flush=True)
for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
    if any(v is not None for v in row[:3]):
        # 日付っぽい行だけ表示
        v0 = row[0]
        if v0 and ("2026" in str(v0) or "46" in str(v0)):
            print(f"行{i}: {row[:8]}", flush=True)
