# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

print("=== TERRA 稼働中全員（サイト・実入り） ===", flush=True)
ws_t = wb["TERRA"]
for row in ws_t.iter_rows(min_row=4, values_only=True):
    if row[2] and "稼働中" in str(row[2]):
        tanto = row[0] or ""
        kubun = row[1] or ""
        name = row[3] or ""
        site = row[8]
        terra_req = row[15]
        okamoto_out = row[16]
        jissyu = row[17]
        shiire = row[13]
        tanka = row[7]
        print(
            f"  {name} | {tanto}{kubun} | サイト{site} | TERRA請求:{terra_req} | 岡本払:{okamoto_out} | 実入り:{jissyu} | 案件:{tanka} 仕入:{shiire}",
            flush=True,
        )

print("\n=== GL 稼働中 ===", flush=True)
ws_gl = wb["グレイスライン"]
for row in ws_gl.iter_rows(min_row=4, values_only=True):
    if row[0] and "稼働中" in str(row[0]):
        name = row[1]
        case = row[4]
        tanka = row[5]
        shiire = row[6]
        gross = row[7]
        gl_req = row[8]
        jissyu = row[9]
        site = row[10]
        kihara = row[14]
        print(
            f"  {name} | サイト{site} | 案件:{tanka} 仕入:{shiire} 粗利:{gross} GL請求:{gl_req} 実入り:{jissyu} 木原:{kihara}",
            flush=True,
        )

print("\n=== FT 稼働中 ===", flush=True)
ws_ft = wb["フラップテック"]
for row in ws_ft.iter_rows(min_row=4, values_only=True):
    if row[1] and "稼働中" in str(row[1]):
        tanto = row[0] or ""
        name = row[2]
        tanka = row[6]
        shiire = row[7]
        gross = row[8]
        ft_req = row[9]
        okamoto_out = row[10]
        jissyu = row[11]
        site = row[12]
        kihara = row[15]
        print(
            f"  {name}({tanto}) | サイト{site} | 案件:{tanka} 仕入:{shiire} 粗利:{gross} FT請求:{ft_req} 岡本払:{okamoto_out} 実入:{jissyu} 木原:{kihara}",
            flush=True,
        )

print("\n=== 直契約 ===", flush=True)
ws_dc = wb["直契約"]
for row in ws_dc.iter_rows(min_row=4, values_only=True):
    if row[0] and "稼働中" in str(row[0]):
        name = row[1]
        req = row[2]
        site = row[3]
        ok_out = row[4]
        jissyu = row[5]
        start = row[6]
        print(f"  {name} | {site} | 請求:{req} 岡本払:{ok_out} 実入:{jissyu} 開始:{start}", flush=True)
