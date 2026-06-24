# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_ft = wb["フラップテック"]

# FTシートの最終データ行を探す
last_data_row = 3
for row in ws_ft.iter_rows(min_row=4):
    if any(cell.value is not None for cell in row):
        last_data_row = row[0].row

new_row = last_data_row + 1
print(f"追加行: {new_row}", flush=True)

# ヘッダー確認（3行目）
headers = [ws_ft.cell(row=3, column=i).value for i in range(1, 20)]
print("ヘッダー:", headers, flush=True)

# データ追加
# 担当, ステータス, 氏名, 参画時期, 期間, 案件/上位, 案件単価, 仕入単価, 粗利, FT請求額, 岡本払出, 実入り, 支払サイト, 勤怠表フロー, 送付先, 木原分, 更新サイクル, 備考
data = {
    1: "岡本折半",  # 担当
    2: "稼働中",  # ステータス
    3: "遠藤健太",  # 氏名
    4: "2026/7",  # 参画時期
    5: "長期",  # 期間
    6: "スウェル",  # 案件/上位
    7: 600000,  # 案件単価
    8: 550000,  # 仕入単価
    9: 50000,  # 粗利
    10: "=I{r}*0.68".format(r=new_row),  # FT請求額（粗利×68%）
    11: "=J{r}/2".format(r=new_row),  # 岡本払出（折半）
    12: "=J{r}-K{r}".format(r=new_row),  # 実入り
    13: 45,  # 支払サイト
    14: "上位からくる",  # 勤怠表フロー
    15: "",  # 送付先
    16: 0,  # 木原分
    17: "単月更新",  # 更新サイクル
    18: "赤坂見附・リモート併用。上位Rezon。精算140-200h上下割。岡本折半=粗利×68%の50%払出。9月15日入金",  # 備考
}

for col, val in data.items():
    ws_ft.cell(row=new_row, column=col, value=val)

wb.save(XLSX_PATH)
print(f"[DONE] FTシートに遠藤健太を追加（行{new_row}）", flush=True)

# 確認
print("\n=== 追加データ確認 ===", flush=True)
row_vals = [ws_ft.cell(row=new_row, column=i).value for i in range(1, 19)]
for i, (h, v) in enumerate(zip(headers, row_vals), 1):
    if v is not None:
        print(f"  {h}: {v}", flush=True)
