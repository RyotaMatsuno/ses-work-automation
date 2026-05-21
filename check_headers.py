import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx', data_only=True)
print("シート:", wb.sheetnames)
for sn in wb.sheetnames[:3]:
    ws = wb[sn]
    print(f"\n=== {sn} ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v for v in row if v is not None):
            print(f"row{i}: {row[:10]}")
        if i > 4:
            break
