import glob
import os

import openpyxl

downloads = r"C:\Users\ma_py\Downloads"
files = glob.glob(os.path.join(downloads, "*.xlsx"))
print("Excelファイル一覧:")
for f in files:
    print(" ", os.path.basename(f))

# 単価マスターを確認
target = None
for f in files:
    if "単価" in os.path.basename(f) or "マスター" in os.path.basename(f):
        target = f
        break

if target:
    print(f"\n対象ファイル: {os.path.basename(target)}")
    wb = openpyxl.load_workbook(target)
    for sh in wb.sheetnames:
        print(f"\nシート: {sh}")
        ws = wb[sh]
        # ヘッダー行確認
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            print(row)
else:
    print("\n単価マスターファイルが見つかりません")
    # 最新のxlsxを確認
    latest = sorted(files, key=os.path.getmtime, reverse=True)[0]
    print(f"最新ファイル: {os.path.basename(latest)}")
    wb = openpyxl.load_workbook(latest)
    for sh in wb.sheetnames:
        print(f"\nシート: {sh}")
        ws = wb[sh]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            print(row)
