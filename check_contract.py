# -*- coding: utf-8 -*-

import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH)
print("シート一覧:", wb.sheetnames, flush=True)

# 各シートの先頭5行を確認
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n=== {sname} ===")
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        print(row[:5])
