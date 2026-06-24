# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_ft = wb["フラップテック"]

# 最終データ行の次に追加（行14か15あたり）
# 原昌志(行5)は退場済みなので14行目に鶴川を追加
new_row = 14

ws_ft.cell(row=new_row, column=1, value="岡本全額")
ws_ft.cell(row=new_row, column=2, value="稼働中")
ws_ft.cell(row=new_row, column=3, value="鶴川慶三")
ws_ft.cell(row=new_row, column=4, value="2026/6")
ws_ft.cell(row=new_row, column=5, value="長期")
ws_ft.cell(row=new_row, column=6, value="アバンテック")
ws_ft.cell(row=new_row, column=7, value=720000)
ws_ft.cell(row=new_row, column=8, value=650000)
ws_ft.cell(row=new_row, column=9, value=f"=G{new_row}-H{new_row}")
ws_ft.cell(row=new_row, column=10, value=f"=I{new_row}*0.68")
ws_ft.cell(row=new_row, column=11, value=f"=J{new_row}")  # 岡本全額払出
ws_ft.cell(row=new_row, column=12, value=0)  # 松野実入り0
ws_ft.cell(row=new_row, column=13, value=45)
ws_ft.cell(row=new_row, column=14, value="上位からくる")
ws_ft.cell(row=new_row, column=15, value=None)
ws_ft.cell(row=new_row, column=16, value=0)
ws_ft.cell(row=new_row, column=17, value="単月更新")
ws_ft.cell(
    row=new_row,
    column=18,
    value="岡本全額払出。粗利7万×68%=47,600円を全額岡本へ払出。松野実入り0円。6月入場→8月15日初回入金",
)

wb.save(XLSX_PATH)
print(f"[OK] 鶴川慶三 FT行{new_row}に追加完了", flush=True)
print("[DONE]", flush=True)
