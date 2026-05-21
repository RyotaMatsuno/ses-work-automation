
import openpyxl, json, glob, os

BAK_DIR = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract"
baks = sorted(glob.glob(BAK_DIR + r"\*bak*.xlsx"))
print("バックアップ一覧:")
for b in baks:
    print(" ", os.path.basename(b))

# 最初のバックアップ（今日の操作前）で鶴川・斎藤を確認
if baks:
    first_bak = baks[0]
    print(f"\n最古バックアップ: {os.path.basename(first_bak)}")
    wb = openpyxl.load_workbook(first_bak, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for cell in row:
                if cell and any(k in str(cell) for k in ["斎藤","齋藤","鶴川","saito","tsurukawa"]):
                    print(f"  sheet={sname} 行{i}: {[str(v)[:15] if v else None for v in row[:6]]}")
