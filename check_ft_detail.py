# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_ft = wb["フラップテック"]

# FTシートのヘッダーと各行を確認
print("=== FTシート全行（稼働中・入場前）===", flush=True)
headers = [ws_ft.cell(row=3, column=i).value for i in range(1, 22)]
for row in ws_ft.iter_rows(min_row=4, values_only=True):
    if row[1] and row[1] in ["稼働中", "入場前"]:
        name = row[2]
        tanto = row[0] or ""
        budget = row[6]  # 案件単価
        cost = row[7]  # 仕入単価
        gross = row[8]  # 粗利
        ft_req = row[9]  # FT請求額
        okamoto = row[10]  # 岡本払出
        jissyu = row[11]  # 実入り
        kihara = row[15]  # 木原分
        print(
            f"{name}({tanto}): 案件{budget} 仕入{cost} 粗利{gross} FT請求{ft_req} 岡本払出{okamoto} 実入り{jissyu} 木原{kihara}",
            flush=True,
        )
