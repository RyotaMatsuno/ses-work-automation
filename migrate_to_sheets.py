
"""
Excel契約マスター → Googleスプレッドシートへ移行
"""
import gspread
from google.oauth2.service_account import Credentials
import openpyxl

CREDS_PATH  = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\google_credentials.json"
EXCEL_PATH  = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
SCOPES = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

creds = Credentials.from_service_account_file(CREDS_PATH, scopes=SCOPES)
gc = gspread.authorize(creds)

# 新規スプレッドシート作成
ss = gc.create("契約マスター_TERRA")
print(f"作成: {ss.title}")
print(f"URL: https://docs.google.com/spreadsheets/d/{ss.id}")

# Excelを読み込んで各シートをコピー
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

for i, sheet_name in enumerate(wb.sheetnames):
    ws_excel = wb[sheet_name]
    rows = list(ws_excel.iter_rows(values_only=True))

    if i == 0:
        # 最初のシートは既存シートを使う
        gs = ss.sheet1
        gs.update_title(sheet_name)
    else:
        gs = ss.add_worksheet(title=sheet_name, rows=len(rows)+10, cols=30)

    # データを文字列に変換してバッチ書き込み
    data = []
    for row in rows:
        data.append([str(v) if v is not None else "" for v in row])

    if data:
        gs.update(data, "A1")
    print(f"  シート「{sheet_name}」: {len(rows)}行 書き込み完了")

# Service Accountのメールに共有は不要（オーナーがService Accountなので松野さんのアカウントに共有する）
# 松野さんのGmailに共有
ss.share("r-matsuno@terra-ltd.co.jp", perm_type="user", role="writer", notify=False)
print(f"\n松野アドレスに共有完了")
print(f"スプレッドシートID: {ss.id}")

# IDをファイルに保存
with open(r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\spreadsheet_id.txt", "w") as f:
    f.write(ss.id)
print("IDを spreadsheet_id.txt に保存")
