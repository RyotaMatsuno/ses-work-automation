# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import shutil
from datetime import datetime

import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
BAK = XLSX_PATH.replace(".xlsx", f"_bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
shutil.copy2(XLSX_PATH, BAK)
print(f"バックアップ: {BAK}", flush=True)

wb = openpyxl.load_workbook(XLSX_PATH)

# ============================================================
# 列マップ（4行目ヘッダー確認済み）
# A=1:担当 B=2:区分 C=3:ステータス D=4:氏名 E=5:参画 F=6:期間
# G=7:案件会社 H=8:単価 I=9:サイト J=10:勤怠 K=11:更新 L=12:業務
# M=13:仕入先 N=14:仕入単価 O=15:粗利 P=16:TERRA請求 Q=17:岡本払出
# R=18:実入り S=19:備考
# ============================================================

ws_t = wb["TERRA"]
terra_map = {}
for i, row in enumerate(ws_t.iter_rows(min_row=5, values_only=True), start=5):
    if row[3]:
        terra_map[str(row[3])] = i


def t(name, col, val):
    r = terra_map.get(name)
    if r:
        ws_t.cell(row=r, column=col, value=val)


# ============================================================
# TERRA プロパー 松野担当（15名）
# 実入り = TERRA請求 × 1.1 × (1 - 0.1021) = 請求×0.9879
# 15,000 × 1.1 = 16,500 - 源泉1,531 = 14,969円
# ============================================================
matsuno_propre = [
    "仲山雄輝",
    "吉田創志",
    "蒲池佑萌",
    "大野稔貴",
    "白須雄太",
    "沼田航陽",
    "赤木",
    "坪井",
    "中村",
    "日比野",
    "永野",
    "安江",
    "相川",
    "富永",
]
for name in matsuno_propre:
    r = terra_map.get(name)
    if not r:
        continue
    ws_t.cell(row=r, column=17, value=0)  # 岡本払出なし
    ws_t.cell(row=r, column=18, value=f"=P{r}*1.1-P{r}*0.1021")  # 実入り（源泉後）
    ws_t.cell(row=r, column=19, value="松野担当 TERRAプロパー。実入り=請求×1.1-源泉(×10.21%)")
    print(f"  [OK] {name}: 実入り数式・備考入力", flush=True)

# 魚谷（退場済み）
t("魚谷", 19, "退場済み（4月末）")

# ============================================================
# TERRA BP折半（松野）
# 実入り = 粗利×50% → TERRA請求 → 税込源泉後
# ============================================================
bp_matsuno = {
    "森": "粗利6万×50%=30,000請求。実入り=請求×1.1-源泉",
    "芹澤": "粗利9万×50%=45,000請求。実入り=請求×1.1-源泉",
    "小山内": "粗利6万×50%=30,000請求。実入り=請求×1.1-源泉。7〜9月延長確定",
}
for name, note in bp_matsuno.items():
    r = terra_map.get(name)
    if not r:
        continue
    ws_t.cell(row=r, column=15, value=f"=H{r}-N{r}")  # 粗利
    ws_t.cell(row=r, column=16, value=f"=O{r}*0.5")  # TERRA請求=粗利×50%
    ws_t.cell(row=r, column=17, value=0)
    ws_t.cell(row=r, column=18, value=f"=P{r}*1.1-P{r}*0.1021")  # 実入り
    ws_t.cell(row=r, column=19, value=note)
    print(f"  [OK] {name}: BP折半数式・備考入力", flush=True)

# 大本（5月末終了）
r = terra_map.get("大本")
if r:
    ws_t.cell(row=r, column=15, value=f"=H{r}-N{r}")
    ws_t.cell(row=r, column=16, value=f"=O{r}*0.5")
    ws_t.cell(row=r, column=17, value=0)
    ws_t.cell(row=r, column=18, value=f"=P{r}*1.1-P{r}*0.1021")
    ws_t.cell(row=r, column=19, value="5月末終了。TERRA折半 粗利×50%")
    print("  [OK] 大本: 5月末終了・数式入力", flush=True)

# ============================================================
# TERRA 岡本担当（天野・岩瀬・加藤T・齋藤）
# TERRA請求15,000・岡本60%=9,000・松野40%=6,000
# 松野実入り = 15,000 × 0.4 × 1.1 - 15,000 × 0.4 × 0.1021
#            = 6,000 × 1.1 - 6,000 × 0.1021 = 5,988円
# ============================================================
okamoto_propre = {
    "天野": "comvace。岡本担当。TERRA請求15,000→岡本60%=9,000/松野40%=6,000(税抜)。松野実入り5,988円/月",
    "岩瀬": "岡本担当。TERRA請求15,000→岡本60%/松野40%。松野実入り5,988円/月",
    "加藤(T)": "岡本担当。TERRA請求15,000→岡本60%/松野40%。松野実入り5,988円/月",
    "齋藤よしまさ": "岡本担当。TERRA請求15,000→岡本60%/松野40%。松野実入り5,988円/月。初月5/15入場0.61人月→7月入金3,653円。6月〜フル5,988円/月",
}
for name, note in okamoto_propre.items():
    r = terra_map.get(name)
    if not r:
        continue
    ws_t.cell(row=r, column=16, value=15000)
    ws_t.cell(row=r, column=17, value=9000)  # 岡本払出60%
    ws_t.cell(row=r, column=18, value=f"=(P{r}-Q{r})*1.1-(P{r}-Q{r})*0.1021")  # 松野実入り（40%分の源泉後）
    ws_t.cell(row=r, column=19, value=note)
    print(f"  [OK] {name}: 岡本担当数式・備考入力", flush=True)

# ============================================================
# TERRA 岡本折半BP（佐々木）
# 粗利×80%がTERRA請求、岡本払出は半分、松野実入りは残り半分の源泉後
# ============================================================
r = terra_map.get("佐々木")
if r:
    ws_t.cell(row=r, column=15, value=f"=H{r}-N{r}")  # 粗利
    ws_t.cell(row=r, column=16, value=f"=O{r}*0.8")  # TERRA請求=粗利×80%
    ws_t.cell(row=r, column=17, value=f"=P{r}*0.5")  # 岡本払出=半分
    ws_t.cell(row=r, column=18, value=f"=(P{r}/2)*1.1-(P{r}/2)*0.1021")  # 松野実入り
    ws_t.cell(
        row=r, column=19, value="岡本折半BP。粗利5万×80%=40,000請求。岡本50%払出。松野実入り=半分の源泉後≒19,757円/月"
    )
    print("  [OK] 佐々木: 岡本折半BP数式入力", flush=True)

# ============================================================
# TERRA 請求なし（GL/FT経由）
# ============================================================
terra_nashi = {
    "山内清": "GL経由。TERRAへの請求なし。GLシートで管理",
    "田中みさ": "GL/FT経由。TERRAへの請求なし。FTシートで管理",
    "木村勇太": "FT経由。TERRAへの請求なし。FTシートで管理",
    "川崎(新)": "FT経由。TERRAへの請求なし。FTシートで管理",
    "橋詰(新)": "TERRAプロパー。単価45万サイト45日。TERRA請求15,000円。松野実入り14,969円/月（源泉後）。4月稼働。",
    "佐々木(入)": "TERRAプロパー扱い。単価61万サイト45日。TERRA請求15,000円確認済み。",
}
for name, note in terra_nashi.items():
    r = terra_map.get(name)
    if not r:
        continue
    ws_t.cell(row=r, column=19, value=note)
    print(f"  [OK] {name}: 備考入力", flush=True)

# 橋詰は実入り数式も入れる
r = terra_map.get("橋詰(新)")
if r:
    ws_t.cell(row=r, column=17, value=0)
    ws_t.cell(row=r, column=18, value=f"=P{r}*1.1-P{r}*0.1021")
    print("  [OK] 橋詰(新): 実入り数式追加", flush=True)

print("[OK] TERRAシート完了", flush=True)

# ============================================================
# FTシート 備考を全員詳細入力
# ============================================================
ws_ft = wb["フラップテック"]
ft_map = {}
for i, row in enumerate(ws_ft.iter_rows(min_row=4, values_only=True), start=4):
    if row[2]:
        ft_map[str(row[2])] = i

ft_notes = {
    "笠井健太": "通常。粗利10万×68%=68,000-木原10,000=実入り58,000(税抜)→税込63,800円。スウェル松堂さん。通信会社運用門前仲町。四半期更新",
    "木村勇太": "通常。粗利7万×68%=47,600-木原10,000=実入り37,600(税抜)→税込41,360円。スウェル。Java赤坂見附。単月更新",
    "加藤": "小坂折半。粗利8万×48%=38,400-木原11,000=実入り27,400(税抜)→税込30,140円。スウェル小坂さん。四半期更新",
    "川崎健太": "通常。粗利4万×68%=27,200-木原5,000=実入り22,200(税抜)→税込24,420円。スウェル。赤坂見附リモート。単月更新",
    "田中みさ": "通常。粗利3万×68%=20,400-木原0=実入り20,400(税抜)→税込22,440円。相馬さん。固定精算。3か月更新",
    "立野和紀": "通常。粗利7万×68%=47,600-木原10,000=実入り37,600(税抜)→税込41,360円。スウェル(エンジニアのミカタ)。食品業界。3か月更新",
    "佐々木駿": "通常。粗利4万×68%=27,200-木原20,000=実入り7,200(税抜)→税込7,920円。スウェル。Java赤坂見附。単月更新",
    "橋本奈緒": "岡本折半。粗利7万×68%=47,600÷2-木原10,000=実入り13,800(税抜)→税込15,180円。スウェル岡本テラシーク。赤坂見附週3出社。単月更新",
    "鶴川慶三": "岡本全額払出。粗利7万×68%=47,600円を全額岡本払出。松野実入り0円。アバンテック。サイト45日。6月入場→8月15日初回入金",
    "吉田祥平": "小坂折半。粗利5万×48%=24,000-木原0=実入り24,000(税抜)→税込26,400円。省庁NW/ICT。FP単価75万→FP-TERRA間70万。精算140-180h中間割。6月入場→8月入金",
    "遠藤健太": "岡本折半。粗利5万×68%=34,000÷2-木原0=松野実入り17,000(税抜)→税込18,700円。スウェル/Rezon。赤坂見附リモート。精算140-200h上下割。7月入場→9月15日初回入金",
}

for name, note in ft_notes.items():
    r = ft_map.get(name)
    if not r:
        print(f"  FT[{name}] NOT FOUND", flush=True)
        continue
    ws_ft.cell(row=r, column=18, value=note)
    print(f"  [OK] FT[{name}]: 備考更新", flush=True)

print("[OK] FTシート完了", flush=True)

# ============================================================
# GLシート 備考入力
# ============================================================
ws_gl = wb["グレイスライン"]
gl_map = {}
for i, row in enumerate(ws_gl.iter_rows(min_row=4, values_only=True), start=4):
    if row[1]:
        gl_map[str(row[1])] = i

gl_notes = {
    "石崎春光": "スウェル(松崎様)。粗利4万×60%=24,000-木原0=実入り24,000(税抜)→税込26,400円。サイト30日（翌月末入金）。勤怠:support_hajimariworks@hajimari.inc",
    "山内清": "スウェル。粗利7万×60%=42,000-木原10,000=実入り32,000(税抜)→税込35,200円。サイト45日。TERRAにも請求なしで入力済み",
    "荒井大輝": "スウェル(テックビズ)。粗利7万×60%=42,000-木原10,000=実入り32,000(税抜)→税込35,200円。サイト45日",
}

for name, note in gl_notes.items():
    r = gl_map.get(name)
    if not r:
        print(f"  GL[{name}] NOT FOUND", flush=True)
        continue
    # GLの備考列を確認して入力
    ws_gl.cell(row=r, column=16, value=note)
    print(f"  [OK] GL[{name}]: 備考更新", flush=True)

print("[OK] GLシート完了", flush=True)

# ============================================================
# 月次サマリーシートに最終確定値を追記
# ============================================================
ws_sum = wb["月次サマリー"]
last_row = ws_sum.max_row

from openpyxl.styles import Font, PatternFill

ws_sum.cell(row=last_row + 2, column=1, value="■ 橋詰・齋藤追加後 最終確定値（2026-06-01）").font = Font(bold=True)
headers = ["入金月", "手取り(税込)", "TERRA源泉", "額面", "前月比", "備考"]
for i, h in enumerate(headers, 1):
    c = ws_sum.cell(row=last_row + 3, column=i, value=h)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", start_color="D9D9D9")

final_data = [
    ("6月入金", 845092, 50028, 895120, "—", "橋詰4月稼働追加(+14,969)"),
    ("7月入金", 901105, 50028, 951133, "+56,013", "橋詰+14,969・齋藤初月+3,653"),
    ("8月入金", 898056, 55133, 953189, "-3,049", "鶴川払出+・吉田追加・木原減・橋詰+14,969・齋藤フル+5,988"),
    ("9月入金", 898056, 55133, 953189, "±0", "遠藤健太は相殺のため変化なし"),
]
for j, row_data in enumerate(final_data):
    for k, val in enumerate(row_data, 1):
        ws_sum.cell(row=last_row + 4 + j, column=k, value=val)

print("[OK] 月次サマリー最終確定値追記", flush=True)

# ============================================================
# 保存
# ============================================================
wb.save(XLSX_PATH)
print("\n[DONE] 保存完了", flush=True)
