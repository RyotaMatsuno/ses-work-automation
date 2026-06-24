# -*- coding: utf-8 -*-
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

XLSX_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH)
ws_t = wb["TERRA"]

for i, row in enumerate(ws_t.iter_rows(min_row=4, values_only=True), start=4):
    if row[3] == "橋詰(新)":
        ws_t.cell(row=i, column=18, value=None)
        print("[OK] 橋詰(新) 備考メモ削除", flush=True)
        break

wb.save(XLSX_PATH)
print("[DONE]", flush=True)
