import openpyxl, sys, shutil
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\ma_py\OneDrive\デスクトップ\ses_work\contract\契約マスター_v6.xlsx"
BACKUP_PATH = EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")\

# バックアップ
shutil.copy2(EXCEL_PATH, BACKUP_PATH)
print(f"バックアップ: {BACKUP_PATH}")

wb = openpyxl.load_workbook(EXCEL_PATH)

# ===== 1. FT: 岡本行（行13=Excel行15、名前なしのダミー行）削除 =====
ws_ft = wb["フラップテック"]
target_row = None
for i, row in enumerate(ws_ft.iter_rows(values_only=True), start=1):
    if row[0] == "岡本" and row[1] is None and row[2] is None:
        target_row = i
        break

if target_row:
    ws_ft.delete_rows(target_row)
    print(f"[FT] 行{target_row}（岡本ダミー行）削除完了")
else:
    print("[FT] 削除対象行が見つかりませんでした")

# ===== 2. TERRA: 2026/4入場人員を「入場前」→「稼働中」に変更 =====
ws_tr = wb["TERRA"]
changed = []
for row in ws_tr.iter_rows():
    vals = [c.value for c in row]
    # ステータスはcol[2]（0-indexed=col C）
    status_cell = row[2]  # col C = ステータス
    sankai_cell = row[4]  # col E = 参画時期
    if status_cell.value == "入場前" and str(sankai_cell.value or "").startswith("2026/4"):
        name = row[3].value  # 氏名
        status_cell.value = "稼働中"
        changed.append(name)

print(f"[TERRA] 稼働中に変更: {changed}")

# ===== 入金予測シートの現状確認 =====
ws_ny = wb["入金予測"]
print("\n=== 入金予測シート現状 ===")
for i, row in enumerate(ws_ny.iter_rows(values_only=True)):
    if any(row):
        print(f"  行{i}: {row[:12]}")

wb.save(EXCEL_PATH)
print("\nExcel保存完了")
